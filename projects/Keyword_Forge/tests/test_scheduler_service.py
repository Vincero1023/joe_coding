from __future__ import annotations

import time
from datetime import datetime, time as dt_time, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.scheduler.service import (
    _JOB_STATUS_COMPLETED,
    JobSchedulerService,
    reset_job_scheduler_service_for_tests,
)


def _fake_pipeline_runner(input_data: dict) -> dict:
    collector = input_data.get("collector") if isinstance(input_data.get("collector"), dict) else {}
    seed = str(collector.get("seed_input") or "").strip()
    category = str(collector.get("category") or "").strip()
    label = seed or category or "default"

    return {
        "collected_keywords": [
            {
                "keyword": label,
                "category": category or "manual",
                "source": "queue_test",
                "raw": label,
            }
        ],
        "expanded_keywords": [
            {
                "keyword": f"{label} 확장",
                "origin": label,
                "type": "queue_test",
            }
        ],
        "analyzed_keywords": [
            {
                "keyword": f"{label} 분석",
                "score": 77.0,
                "profitability_grade": "A",
                "attackability_grade": "2",
            }
        ],
        "selected_keywords": [
            {
                "keyword": f"{label} 선별",
                "score": 71.0,
                "profitability_grade": "A",
                "attackability_grade": "2",
            }
        ],
        "longtail_suggestions": [
            {
                "longtail_keyword": f"{label} 롱테일",
                "source_keyword": label,
            }
        ],
        "generated_titles": [
            {
                "keyword": f"{label} 선별",
                "target_mode": "single",
                "source_kind": "selected_keyword",
                "titles": {
                    "naver_home": [f"{label} 홈1", f"{label} 홈2"],
                    "blog": [f"{label} 블로그1", f"{label} 블로그2"],
                },
                "quality_report": {
                    "label": "양호",
                    "bundle_score": 88,
                },
            }
        ],
    }


def _wait_for_job(service: JobSchedulerService, job_id: str, *, timeout_seconds: float = 5.0) -> dict:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        snapshot = service.get_snapshot()
        job = next(item for item in snapshot["jobs"] if item["job_id"] == job_id)
        if job["status"] in {"completed", "partial", "failed", "blocked", "canceled"}:
            return job
        time.sleep(0.05)
    raise AssertionError("queue job did not finish in time")


def test_seed_batch_job_runs_and_writes_xlsx(tmp_path: Path) -> None:
    state_path = tmp_path / "scheduler_state.json"
    output_dir = tmp_path / "exports"
    service = reset_job_scheduler_service_for_tests(
        state_path=state_path,
        output_dir=output_dir,
        runner=_fake_pipeline_runner,
        poll_interval_seconds=0.1,
    )

    try:
        snapshot = service.enqueue_seed_batch_job(
            name="보험 배치",
            seeds=["보험 추천", "대출 비교"],
            base_input={"title_options": {"mode": "template"}},
        )

        job_id = snapshot["jobs"][0]["job_id"]
        job = _wait_for_job(service, job_id)

        assert job["status"] == _JOB_STATUS_COMPLETED
        assert job["completed_count"] == 2
        assert Path(job["artifact_path"]).exists()

        workbook = load_workbook(job["artifact_path"])
        assert "summary" in workbook.sheetnames
        assert "titles" in workbook.sheetnames
        assert workbook["summary"].max_row >= 3
        assert workbook["titles"].max_row >= 3
    finally:
        service.shutdown()


def test_completed_seed_batch_snapshot_preserves_seed_values(tmp_path: Path) -> None:
    state_path = tmp_path / "scheduler_state.json"
    output_dir = tmp_path / "exports"
    service = reset_job_scheduler_service_for_tests(
        state_path=state_path,
        output_dir=output_dir,
        runner=_fake_pipeline_runner,
        poll_interval_seconds=0.1,
    )

    try:
        snapshot = service.enqueue_seed_batch_job(
            name="검증 시드 배치",
            seeds=["로지텍 마우스", "오사카 가성비 호텔", "닌텐도 스위치2 사전예약"],
            base_input={"title_options": {"mode": "template"}},
        )

        job_id = snapshot["jobs"][0]["job_id"]
        _wait_for_job(service, job_id)
        latest_snapshot = service.get_snapshot()
        job = next(item for item in latest_snapshot["jobs"] if item["job_id"] == job_id)

        assert [item["value"] for item in job["items"]] == [
            "로지텍 마우스",
            "오사카 가성비 호텔",
            "닌텐도 스위치2 사전예약",
        ]
        assert all(item["status"] == _JOB_STATUS_COMPLETED for item in job["items"])
    finally:
        service.shutdown()


def test_daily_category_routine_enqueues_due_job(tmp_path: Path) -> None:
    state_path = tmp_path / "scheduler_state.json"
    output_dir = tmp_path / "exports"
    service = reset_job_scheduler_service_for_tests(
        state_path=state_path,
        output_dir=output_dir,
        runner=_fake_pipeline_runner,
        poll_interval_seconds=0.1,
    )

    reference_now = datetime(2030, 3, 23, 5, 50)

    try:
        service.create_daily_category_routine(
            name="아침 카테고리 루틴",
            categories=["비즈니스·경제", "IT·컴퓨터"],
            time_of_day=dt_time(hour=6, minute=0),
            weekdays=[5],
            base_input={"title_options": {"mode": "template"}},
            reference_now=reference_now,
        )

        created_count = service.queue_due_routines(now=reference_now + timedelta(minutes=15))
        assert created_count == 1

        snapshot = service.get_snapshot()
        routine_jobs = [job for job in snapshot["jobs"] if job["source"] == "routine"]
        assert len(routine_jobs) == 1
        assert routine_jobs[0]["item_count"] == 2
        assert routine_jobs[0]["scheduled_for"]
    finally:
        service.shutdown()
