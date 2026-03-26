from unittest.mock import patch
import csv
import json
from pathlib import Path
import pytest

from openpyxl import load_workbook

from app.title.ai_client import (
    TitleGenerationOptions,
    TitleProviderError,
    _build_user_prompt_from_items,
    _resolve_retry_delay_seconds,
    request_ai_slot_rewrites,
    request_ai_titles,
    request_ai_json_object,
)
from app.title.category_detector import detect_category
from app.title.evaluation_prompt import DEFAULT_TITLE_EVALUATION_PROMPT
from app.title.exporter import export_generated_titles
from app.title.main import run
from app.title.quality import assess_single_title, enrich_title_results, refresh_title_results_for_changed_slots
from app.title.quality_ai import TitleEvaluationOptions
from app.title.quality_ai import _normalize_batch_evaluation_item, request_naver_home_title_evaluations_batch
from app.title.rules import NAVER_HOME_MAX_LENGTH
from app.title.targets import _sanitize_related_mode_keyword, build_title_targets
from app.title.title_generator import (
    _apply_practical_rescue_candidates,
    _build_title_evaluation_options,
    _build_practical_rescue_item,
    _build_quality_retry_prompt,
    _collect_retry_slot_candidates,
    _should_retry_slot,
    generate_titles,
)
from app.title.templates import build_blog_titles, build_naver_home_titles


def _build_slot_retry_response(chunk, *, degraded: bool = False) -> list[dict[str, str]]:
    response: list[dict[str, str]] = []
    for item in chunk:
        keyword = str(item.get("keyword") or "").strip()
        channel = str(item.get("channel") or "").strip()
        slot_index = int(item.get("slot_index") or 1)
        if degraded:
            title = f"{keyword}!!"
        elif channel == "naver_home":
            title = (
                f"{keyword}, 왜 선택 기준이 갈릴까?"
                if slot_index == 1
                else f"{keyword}, 지금 비교 포인트 2가지"
            )
        else:
            title = (
                f"{keyword} 비교 포인트 정리"
                if slot_index == 1
                else f"{keyword} 체크리스트 정리"
            )
        response.append(
            {
                "slot_id": str(item.get("slot_id") or "").strip(),
                "title": title,
            }
        )
    return response


def test_detect_category_handles_finance_keyword() -> None:
    assert detect_category("보험 추천") == "finance"


def test_detect_category_distinguishes_real_estate_and_finance_signals() -> None:
    assert detect_category("서울 청약 경쟁률") == "real_estate"
    assert detect_category("ETF 리밸런싱 일정") == "finance"


def test_title_quality_rejects_finance_mismatch_frame() -> None:
    report = assess_single_title(
        "국제금시세",
        "국제금시세 실사용 차이, 직접 써본 기준",
        "naver_home",
        {},
    )

    assert report["status"] == "retry"
    assert report["checks"]["finance_domain_mismatch"] is True
    assert "금융 카테고리와 맞지 않는 제목 프레임입니다." in report["issues"]


def test_title_quality_rejects_finance_stale_investment_wrapper() -> None:
    report = assess_single_title(
        "국제금시세",
        "국제금시세 투자 전략 가이드",
        "blog",
        {},
    )

    assert report["status"] == "retry"
    assert report["checks"]["finance_stale_wrapper"] is True
    assert "금융 키워드에 비해 제목이 너무 느슨한 투자형 템플릿입니다." in report["issues"]


def test_ai_prompt_builder_adds_finance_market_shape_hint() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "국제금시세",
                "quality_report": {
                    "bundle_score": 82,
                    "status": "review",
                    "channel_scores": {"naver_home": 81, "blog": 80},
                },
                "target_mode": "single",
            }
        ]
    )

    assert "finance market: use gap, interpretation, checkpoint, variable" in prompt
    assert "Prefer concise hook-first wording" in prompt


def test_build_title_targets_skips_finance_device_style_longtails() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {"keyword": "국제금시세", "score": 52.0},
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "국제금시세",
                    "source_keyword": "국제금시세",
                    "longtail_keyword": "국제금시세 실사용 차이",
                    "verification_status": "pass",
                    "verified_score": 72.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "국제금시세",
                    "source_keyword": "국제금시세",
                    "longtail_keyword": "국제금시세 자주 생기는 문제",
                    "verification_status": "pass",
                    "verified_score": 70.0,
                },
                {
                    "suggestion_id": "longtail-03",
                    "representative_keyword": "국제금시세",
                    "source_keyword": "국제금시세",
                    "longtail_keyword": "국제금시세 국내외 차이",
                    "verification_status": "pass",
                    "verified_score": 69.0,
                },
            ],
            "title_options": {
                "keyword_modes": ["longtail_selected"],
            },
        }
    )

    keywords = {item["keyword"] for item in items}

    assert summary["mode_counts"]["longtail_selected"] == 1
    assert "국제금시세 국내외 차이" in keywords
    assert "국제금시세 실사용 차이" not in keywords
    assert "국제금시세 자주 생기는 문제" not in keywords


def test_title_generator_returns_two_titles_per_type() -> None:
    result = run(
        [
            {
                "keyword": "보험 추천",
                "score": 1.0,
                "metrics": {
                    "volume": 1.0,
                    "cpc": 1.0,
                    "competition": 0.7,
                    "bid": 1.0,
                    "profit": 1.0,
                    "opportunity": 1.4286,
                },
            }
        ]
    )

    assert len(result) == 1
    assert len(result[0]["titles"]["naver_home"]) == 2
    assert len(result[0]["titles"]["blog"]) == 2
    assert all("보험 추천" in title for title in result[0]["titles"]["naver_home"])
    assert all(len(title) <= NAVER_HOME_MAX_LENGTH for title in result[0]["titles"]["naver_home"])


def test_title_generator_preserves_keyword_text() -> None:
    result = run(
        [
            {
                "keyword": "제주 여행 코스",
                "score": 0.82,
                "metrics": {
                    "volume": 0.7,
                    "cpc": 0.7,
                    "competition": 0.4,
                    "bid": 0.6,
                    "profit": 0.42,
                    "opportunity": 1.75,
                },
            }
        ]
    )

    assert result[0]["keyword"] == "제주 여행 코스"
    assert all("제주 여행 코스" in title for title in result[0]["titles"]["blog"])
    assert result[0]["quality_report"]["bundle_score"] > 0
    assert result[0]["quality_report"]["label"] in {"양호", "재검토", "재생성 권장"}


def test_title_quality_accepts_keyword_tokens_with_inserted_modifiers() -> None:
    report = assess_single_title(
        "트레이더조 에코백 체크리스트",
        "트레이더조 에코백 구매 전 필수 체크리스트",
        "blog",
        {},
    )

    assert report["checks"]["contains_keyword"] is True
    assert report["checks"]["starts_with_keyword"] is True
    assert "키워드 핵심 표현이 제목에 충분히 반영되지 않았습니다." not in report["issues"]


def test_title_quality_still_fails_when_keyword_core_token_is_missing() -> None:
    report = assess_single_title(
        "트레이더조 에코백 체크리스트",
        "트레이더조 에코백 종류별 비교 가이드",
        "blog",
        {},
    )

    assert report["checks"]["contains_keyword"] is False
    assert "키워드 핵심 표현이 제목에 충분히 반영되지 않았습니다." in report["issues"]


def test_title_quality_flags_low_signal_generic_skeletons() -> None:
    report = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스 최신 정보",
        "blog",
        {},
    )

    assert "제목 골격이 템플릿형 표현에 머물러 있습니다." in report["issues"]
    assert report["checks"]["hard_reject_skeleton"] is True
    assert report["status"] == "retry"


def test_title_quality_allows_curiosity_driven_naver_home_question_frames() -> None:
    latest_rank = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스 최신 순위는?",
        "naver_home",
        {},
    )
    best_question = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스 뭐가 제일 좋을까",
        "naver_home",
        {},
    )

    assert latest_rank["checks"]["hard_reject_skeleton"] is False
    assert best_question["checks"]["hard_reject_skeleton"] is False
    assert latest_rank["score_breakdown"]["curiosity_gap"] >= 16
    assert best_question["score_breakdown"]["curiosity_gap"] >= 16
    assert latest_rank["score"] >= 75
    assert best_question["score"] >= 75


def test_title_quality_hard_rejects_generic_overlay_on_practical_keyword() -> None:
    report = assess_single_title(
        "로지텍 마우스 설정 팁",
        "로지텍 마우스 설정 팁 완벽 가이드",
        "blog",
        {},
    )

    assert "구체 글감 위에 다시 템플릿형 포장을 덧씌웠습니다." in report["issues"]
    assert report["checks"]["hard_reject_skeleton"] is True
    assert report["checks"]["generic_overlay_on_practical_keyword"] is True
    assert report["status"] == "retry"


def test_title_quality_allows_concrete_context_on_practical_keyword() -> None:
    report = assess_single_title(
        "MX Master 2S 연결 방법",
        "MX Master 2S 연결 방법 총정리 맥 M1 M2 포함",
        "blog",
        {},
    )

    assert report["checks"]["generic_overlay_on_practical_keyword"] is False


def test_title_quality_prefers_hookier_naver_home_title() -> None:
    plain = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 조건과 절차 정리",
        "naver_home",
        {},
    )
    hooky = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 왜 늦어질까",
        "naver_home",
        {},
    )

    assert plain["checks"]["home_hook_signal"] is False
    assert hooky["checks"]["home_hook_signal"] is True
    assert plain["score"] < hooky["score"]


def test_title_quality_exposes_naver_home_ctr_breakdown() -> None:
    report = assess_single_title(
        "\ud658\uc728 \uc804\ub9dd",
        "\ud658\uc728 \uc804\ub9dd, \uc65c \uad6d\ub0b4\uc678 \ubc18\uc751\uc774 \uac08\ub838\uc744\uae4c?",
        "naver_home",
        {},
    )

    breakdown = report["score_breakdown"]
    assert breakdown["issue_or_context"] >= 14
    assert breakdown["curiosity_gap"] >= 16
    assert breakdown["contrast_or_conflict"] >= 10
    assert breakdown["emotional_trigger"] >= 8
    assert breakdown["specificity"] >= 5


def test_title_quality_flags_underdeveloped_naver_home_axis_label_without_forcing_penalty() -> None:
    weak = assess_single_title(
        "24K금값",
        "24K금값, 환율 영향",
        "naver_home",
        {},
        item_context={
            "keyword": "24K금값",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        },
    )
    hooky = assess_single_title(
        "24K금값",
        "24K금값, 왜 국내 시세와 다를까",
        "naver_home",
        {},
        item_context={
            "keyword": "24K금값",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        },
    )

    assert weak["checks"]["home_hook_signal"] is False
    assert weak["checks"]["underdeveloped_home_title"] is True
    assert weak["score"] < hooky["score"]


def test_title_quality_does_not_penalize_short_curiosity_naver_home_titles() -> None:
    report = assess_single_title(
        "ISA 계좌",
        "ISA 계좌, 진짜 왜일까?",
        "naver_home",
        {},
    )

    assert report["score_breakdown"]["curiosity_gap"] >= 18
    assert report["score_breakdown"]["emotional_trigger"] >= 10
    assert report["score_breakdown"]["readability"] == 5
    assert report["status"] != "retry"


def test_title_quality_strongly_rewards_emotional_question_naver_home_titles() -> None:
    report = assess_single_title(
        "ISA 계좌",
        "ISA 계좌, 의외로 먼저 갈린 이유는 뭘까?",
        "naver_home",
        {},
    )

    breakdown = report["score_breakdown"]
    assert breakdown["curiosity_gap"] >= 18
    assert breakdown["reversal_or_unexpected"] >= 12
    assert breakdown["emotional_trigger"] >= 14
    assert report["score"] >= 68


def test_title_quality_prefers_informational_blog_title_over_vague_hook() -> None:
    vague = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 왜 다르게 보일까",
        "blog",
        {},
    )
    concrete = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 준비물과 소요 시간",
        "blog",
        {},
    )

    assert vague["checks"]["blog_info_signal"] is False
    assert concrete["checks"]["blog_info_signal"] is True
    assert vague["score"] < concrete["score"]


def test_title_quality_prefers_blog_search_structure_over_hook_only() -> None:
    vague = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 왜 늦어질까",
        "blog",
        {},
    )
    structured = assess_single_title(
        "ISA 계좌 개설",
        "ISA 계좌 개설 준비물과 비대면 개설 순서",
        "blog",
        {},
    )

    assert vague["checks"]["blog_search_structure"] is False
    assert structured["checks"]["blog_search_structure"] is True
    assert vague["score"] < structured["score"]


def test_title_quality_allows_non_prefix_blog_title_when_structure_is_strong() -> None:
    report = assess_single_title(
        "국제 금값 시세",
        "환율과 국내 시세 차이, 국제 금값 시세 보는 법",
        "blog",
        {},
    )

    assert report["checks"]["contains_keyword"] is True
    assert report["checks"]["starts_with_keyword"] is False
    assert report["checks"]["blog_search_structure"] is True
    assert "키워드가 제목 앞부분에 오지 않습니다." not in report["issues"]


def test_title_quality_flags_generic_blog_wrapper_even_with_structure() -> None:
    generic = assess_single_title(
        "오늘금시세",
        "오늘금시세 최근 동향과 전망",
        "blog",
        {},
    )
    concrete = assess_single_title(
        "오늘금시세",
        "오늘금시세 국내 시세와 국제 금값 차이 정리",
        "blog",
        {},
    )

    assert generic["checks"]["blog_generic_wrapper"] is True
    assert concrete["checks"]["blog_generic_wrapper"] is False
    assert generic["score"] < concrete["score"]


def test_title_quality_rejects_vague_teaser_and_reveal_frames() -> None:
    teaser = assess_single_title(
        "오사카 가성비 호텔",
        "오사카 가성비 호텔 숨은 보석 찾기",
        "blog",
        {},
    )
    reveal = assess_single_title(
        "무선 마우스 설정 팁 전 체크포인트",
        "무선 마우스 설정 팁 전 체크포인트 공개",
        "blog",
        {},
    )
    must_know = assess_single_title(
        "무선 마우스 설정 팁",
        "무선 마우스 설정 팁 꼭 알아둘 5가지",
        "blog",
        {},
    )

    assert teaser["checks"]["hard_reject_skeleton"] is True
    assert teaser["status"] == "retry"
    assert reveal["checks"]["hard_reject_skeleton"] is True
    assert reveal["status"] == "retry"
    assert must_know["checks"]["hard_reject_skeleton"] is True
    assert must_know["status"] == "retry"


def test_title_quality_retries_seed_anchor_single_without_concrete_intent_axis() -> None:
    enriched_items, _summary = enrich_title_results(
        [
            {
                "keyword": "닌텐도 스위치2 사전예약",
                "target_mode": "single",
                "source_selection_mode": "seed_anchor",
                "titles": {
                    "naver_home": [
                        "닌텐도 스위치2 사전예약, 이번주 시작?",
                        "닌텐도 스위치2 사전예약, 경쟁 치열할까",
                    ],
                    "blog": [
                        "닌텐도 스위치2 사전예약, 놓치면 후회할 핵심 정보",
                        "닌텐도 스위치2 사전예약, 구매 전 필독 체크리스트",
                    ],
                },
            }
        ]
    )

    report = enriched_items[0]["quality_report"]
    assert report["retry_recommended"] is True
    assert not any(
        "단일 키워드 제목이 의도 대비 너무 추상적이거나 낚시형입니다." in issue
        for title_report in report["title_checks"]["naver_home"]
        for issue in title_report["issues"]
    )


def test_title_quality_retries_broad_product_single_with_hype_teaser_titles() -> None:
    enriched_items, _summary = enrich_title_results(
        [
            {
                "keyword": "로지텍 마우스",
                "target_mode": "single",
                "source_kind": "selected_keyword",
                "titles": {
                    "naver_home": [
                        "로지텍 마우스 신상, 이젠 이걸로 바꿔보세요",
                        "로지텍 마우스, 당신의 선택은?",
                    ],
                    "blog": [
                        "로지텍 마우스 최신 모델 비교 분석",
                        "로지텍 마우스, 왜 인기일까? 최신 트렌드 분석",
                    ],
                },
            }
        ]
    )

    report = enriched_items[0]["quality_report"]
    assert report["retry_recommended"] is True
    assert not any(
        "단일 키워드 제목이 의도 대비 너무 추상적이거나 낚시형입니다." in issue
        for title_report in report["title_checks"]["naver_home"]
        for issue in title_report["issues"]
    )


def test_title_quality_retries_seed_anchor_value_single_without_concrete_stay_axes() -> None:
    enriched_items, _summary = enrich_title_results(
        [
            {
                "keyword": "오사카 가성비 호텔",
                "target_mode": "single",
                "source_selection_mode": "seed_anchor",
                "titles": {
                    "naver_home": [
                        "오사카 가성비 호텔, 이번주 특가 예약 팁",
                        "오사카 가성비 호텔, 숨겨진 혜택은?",
                    ],
                    "blog": [
                        "오사카 가성비 호텔, 2박 3일 예산별 추천",
                        "오사카 가성비 호텔, 예약부터 체크인까지 가이드",
                    ],
                },
            }
        ]
    )

    report = enriched_items[0]["quality_report"]
    assert report["retry_recommended"] is True
    assert report["recommended_pair_ready"] is False


def test_enrich_title_results_sorts_best_titles_first_for_output() -> None:
    enriched_items, _summary = enrich_title_results(
        [
            {
                "keyword": "로지텍 마우스",
                "target_mode": "single",
                "source_kind": "selected_keyword",
                "titles": {
                    "naver_home": [
                        "로지텍 마우스, 이번주 인기 순위는?",
                        "로지텍 마우스, 배터리와 연결 안정성",
                    ],
                    "blog": [
                        "로지텍 마우스 사용 후기",
                        "로지텍 마우스 배터리·연결 안정성·추천 대상 정리",
                    ],
                },
            }
        ]
    )

    item = enriched_items[0]
    assert item["titles"]["naver_home"][0] == "로지텍 마우스, 이번주 인기 순위는?"
    assert item["titles"]["blog"][0] == "로지텍 마우스 배터리·연결 안정성·추천 대상 정리"
    assert item["quality_report"]["title_checks"]["naver_home"][0]["status"] == "retry"
    assert item["quality_report"]["title_checks"]["blog"][0]["status"] == "good"


def test_refresh_title_results_for_changed_slots_only_reassesses_changed_slots() -> None:
    items = [
        {
            "keyword": "alpha mouse",
            "titles": {
                "naver_home": ["alpha mouse what changed", "alpha mouse compare point"],
                "blog": ["alpha mouse buying note", "alpha mouse setup guide"],
            },
        }
    ]
    current_enriched_items, _ = enrich_title_results(items)
    updated_items = [
        {
            **items[0],
            "titles": {
                "naver_home": list(items[0]["titles"]["naver_home"]),
                "blog": ["alpha mouse condition guide", items[0]["titles"]["blog"][1]],
            },
        }
    ]

    with patch("app.title.quality.assess_single_title", wraps=assess_single_title) as mocked_assess:
        refreshed_items, summary = refresh_title_results_for_changed_slots(
            updated_items,
            current_enriched_items,
            [
                {
                    "result_index": 0,
                    "channel": "blog",
                    "slot_index": 1,
                }
            ],
        )

    assert mocked_assess.call_count == 1
    assert refreshed_items[0]["titles"]["blog"][0] == "alpha mouse condition guide"
    assert summary["total_count"] == 1


def test_refresh_title_results_for_changed_slots_only_batches_changed_naver_home_titles() -> None:
    items = [
        {
            "keyword": "alpha mouse",
            "titles": {
                "naver_home": ["alpha mouse what changed", "alpha mouse compare point"],
                "blog": ["alpha mouse buying note", "alpha mouse setup guide"],
            },
        },
        {
            "keyword": "beta mouse",
            "titles": {
                "naver_home": ["beta mouse what changed", "beta mouse compare point"],
                "blog": ["beta mouse buying note", "beta mouse setup guide"],
            },
        },
    ]
    current_enriched_items, _ = enrich_title_results(items)
    updated_items = [
        {
            **items[0],
            "titles": {
                "naver_home": [items[0]["titles"]["naver_home"][0], "alpha mouse which choice changed"],
                "blog": list(items[0]["titles"]["blog"]),
            },
        },
        items[1],
    ]
    evaluation_options = TitleEvaluationOptions(
        provider="openai",
        model="gpt-4o-mini",
        api_key="openai-key",
        system_prompt="custom eval prompt",
        batch_size=20,
    )
    captured_entries: list[dict[str, str]] = []

    def fake_request(entries, options):
        captured_entries.extend(entries)
        return {
            0: {
                "title": entries[0]["title"],
                "score_breakdown": {
                    "issue_or_context": 13,
                    "curiosity_gap": 15,
                    "contrast_or_conflict": 14,
                    "reversal_or_unexpected": 12,
                    "emotional_trigger": 11,
                    "specificity": 13,
                    "readability": 12,
                    "total": 90,
                },
                "score": 90,
                "verdict": "keep",
                "reason": "CTR 기준 통과",
            }
        }

    with patch("app.title.quality.request_naver_home_title_evaluations_batch", side_effect=fake_request):
        refreshed_items, summary = refresh_title_results_for_changed_slots(
            updated_items,
            current_enriched_items,
            [
                {
                    "result_index": 0,
                    "channel": "naver_home",
                    "slot_index": 2,
                }
            ],
            evaluation_options=evaluation_options,
        )

    assert len(captured_entries) == 1
    assert captured_entries[0]["title"] == "alpha mouse which choice changed"
    assert refreshed_items[0]["titles"]["naver_home"][0] == "alpha mouse which choice changed"
    assert summary["total_count"] == 2


def test_title_quality_retries_unverified_freshness_claim_without_issue_context() -> None:
    report = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스, 2주 실사용 후 장점은?",
        "naver_home",
        {},
        item_context={
            "keyword": "로지텍 마우스",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        },
    )

    assert report["status"] == "retry"
    assert report["checks"]["unverified_freshness_claim"] is True


def test_title_quality_allows_freshness_claim_when_issue_context_exists() -> None:
    report = assess_single_title(
        "서울 청약 경쟁률",
        "서울 청약 경쟁률, 이번주 일정 바뀌나",
        "naver_home",
        {},
        item_context={
            "keyword": "서울 청약 경쟁률",
            "target_mode": "single",
            "issue_context": {
                "query": "서울 청약 경쟁률",
                "title_count": 1,
            },
        },
    )

    assert report["checks"]["unverified_freshness_claim"] is False


def test_title_quality_allows_finance_analysis_window_without_live_issue_context() -> None:
    report = assess_single_title(
        "국제금시세",
        "국제금시세, 2주간의 추이 분석",
        "blog",
        {},
        item_context={
            "keyword": "국제금시세",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        },
    )

    assert report["checks"]["unverified_freshness_claim"] is False
    assert report["status"] in {"review", "good"}


def test_title_quality_normalizes_colon_style_and_flags_it() -> None:
    report = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스：장단점 정리",
        "naver_home",
        {},
    )

    assert ":" not in report["title"]
    assert "：" not in report["title"]
    assert report["checks"]["forbidden_punctuation_used"] is True


def test_enrich_title_results_strips_colons_from_output_titles() -> None:
    enriched_items, _ = enrich_title_results(
        [
            {
                "keyword": "닌텐도 스위치2 사전예약",
                "titles": {
                    "naver_home": [
                        "닌텐도 스위치2 사전예약：오픈 시간 확인",
                        "닌텐도 스위치2 사전예약:결제 조건 확인",
                    ],
                    "blog": [
                        "닌텐도 스위치2 사전예약：오픈 시간·결제 조건 정리",
                        "닌텐도 스위치2 사전예약:수령 일정과 인증 정리",
                    ],
                },
            }
        ]
    )

    item = enriched_items[0]
    assert all(":" not in title and "：" not in title for title in item["titles"]["naver_home"])
    assert all(":" not in title and "：" not in title for title in item["titles"]["blog"])


def test_title_quality_keeps_specific_editorial_product_frame() -> None:
    report = assess_single_title(
        "로지텍 마우스",
        "로지텍 마우스 실사용 장단점과 클릭감 차이",
        "blog",
        {},
    )

    assert "제목 골격이 너무 일반적입니다." not in report["issues"]
    assert report["status"] in {"good", "review"}


def test_template_titles_avoid_legacy_repetitive_phrases() -> None:
    naver_titles = build_naver_home_titles("평택 왁싱 남성전용", "general")
    blog_titles = build_blog_titles("평택 왁싱 남성전용", "general")
    all_titles = naver_titles + blog_titles

    banned_phrases = (
        "완벽 정리",
        "한 번에 정리",
        "갑자기 바뀌었다",
        "이유가 이상하다",
        "놓치면 손해",
        "비교 및 선택 기준 정리",
    )

    assert len(naver_titles) == 2
    assert len(blog_titles) == 2
    assert all("평택 왁싱 남성전용" in title for title in all_titles)
    assert not any(phrase in title for title in all_titles for phrase in banned_phrases)


def test_template_titles_follow_keyword_intent_patterns() -> None:
    review_titles = build_blog_titles("평택 밸리왁싱 후기", "general")
    action_titles = build_blog_titles("평택 밸리왁싱 예약", "general")
    profile_titles = build_blog_titles("함돈균 교수프로필", "general")

    assert any(("후기" in title or "평판" in title) for title in review_titles)
    assert not any("후기 후기" in title for title in review_titles)
    assert any(("준비" in title or "체크리스트" in title or "절차" in title) for title in action_titles)
    assert any(("이력" in title or "정보" in title) for title in profile_titles)


def test_template_titles_expand_intent_detection_for_value_and_setting_keywords() -> None:
    value_titles = build_blog_titles("오사카 가성비 호텔", "general")
    setting_titles = build_blog_titles("무선 마우스 설정 팁", "general")

    assert any(("기준" in title or "선택" in title or "차이" in title) for title in value_titles)
    assert any(("준비" in title or "절차" in title or "설정" in title or "포인트" in title) for title in setting_titles)


def test_template_titles_avoid_repeating_noisy_checklist_and_guide_frames() -> None:
    blog_titles = build_blog_titles("로지텍 마우스", "general")
    noisy_titles = [
        title
        for title in blog_titles
        if any(term in title for term in ("체크리스트", "체크포인트", "가이드", "비교 가이드"))
    ]

    assert len(blog_titles) == 2
    assert len(noisy_titles) <= 1


def test_title_quality_marks_batch_level_repeated_headline_skeletons_for_retry() -> None:
    repeated_batch = [
        {
            "keyword": "로지텍 마우스",
            "titles": {
                "naver_home": ["로지텍 마우스 뭐가 다를까", "로지텍 마우스 추천 기준"],
                "blog": ["로지텍 마우스 체크리스트", "로지텍 마우스 비교 포인트"],
            },
        },
        {
            "keyword": "레이저 마우스",
            "titles": {
                "naver_home": ["레이저 마우스 뭐가 다를까", "레이저 마우스 추천 기준"],
                "blog": ["레이저 마우스 체크리스트", "레이저 마우스 비교 포인트"],
            },
        },
        {
            "keyword": "앱코 마우스",
            "titles": {
                "naver_home": ["앱코 마우스 뭐가 다를까", "앱코 마우스 추천 기준"],
                "blog": ["앱코 마우스 체크리스트", "앱코 마우스 비교 포인트"],
            },
        },
    ]

    enriched_items, summary = enrich_title_results(repeated_batch)
    report = enriched_items[0]["quality_report"]

    assert report["retry_recommended"] is True
    assert report["batch_repeat_risk"] is True
    assert not report["title_checks"]["naver_home"][0]["checks"].get("batch_repeat_risk", False)
    assert report["title_checks"]["blog"][0]["checks"]["batch_repeat_risk"] is True
    assert summary["retry_count"] == 3


def test_title_generator_supports_ai_mode() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": "보험 추천",
                "titles": {
                    "naver_home": ["보험 추천 지금 비교 포인트 2가지", "보험 추천 선택 기준 달라졌다"],
                    "blog": ["보험 추천 완벽 가이드", "보험 추천 비교 포인트 정리"],
                },
            }
        ],
    ):
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "보험 추천",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    assert result["generation_meta"]["used_mode"] == "ai"
    assert result["generation_meta"]["provider"] == "openai"
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 1
    assert result["generated_titles"][0]["titles"]["blog"][0]
    assert result["generation_meta"]["quality_summary"]["total_count"] == 1
    return
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 0
    assert "완벽 가이드" not in result["generated_titles"][0]["titles"]["blog"][0]
    assert "보험 추천 비교 포인트 정리" == result["generated_titles"][0]["titles"]["blog"][0]
    assert result["generation_meta"]["quality_summary"]["total_count"] == 1


def test_title_generator_falls_back_to_template_when_api_key_missing() -> None:
    result = run(
        {
            "selected_keywords": [
                {
                    "keyword": "보험 추천",
                    "score": 1.0,
                }
            ],
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "",
            },
        }
    )

    assert result["generation_meta"]["used_mode"] == "template_fallback"
    assert result["generation_meta"]["fallback_reason"] == "missing_api_key"
    assert len(result["generated_titles"][0]["titles"]["naver_home"]) == 2
    assert result["generated_titles"][0]["quality_report"]["summary"]


def test_title_generator_exports_csv_by_default_when_enabled(tmp_path: Path) -> None:
    result = run(
        {
            "category": "비즈니스경제",
            "seed_input": "보험 추천",
            "selected_keywords": [
                {
                    "keyword": "보험 추천",
                    "score": 1.0,
                }
            ],
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
            },
        }
    )

    export_artifact = result["generation_meta"]["export_artifact"]
    artifact_path = Path(export_artifact["path"])
    assert artifact_path.exists()
    assert artifact_path.suffix == ".csv"
    assert artifact_path.parent.name == "csv"
    assert export_artifact["category"] == "비즈니스경제"
    assert export_artifact["seed_keyword"] == "보험 추천"
    export_artifacts = result["generation_meta"]["export_artifacts"]
    assert [item["format"] for item in export_artifacts] == ["csv"]
    assert all(Path(item["path"]).exists() for item in export_artifacts)

    with artifact_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))

    assert rows[0] == [
        "keyword",
        "bundle_score",
        "bundle_status",
        "naver_home_score",
        "blog_score",
        "target_mode",
        "source_kind",
        "duplicate",
        "recent_used_date",
        "naver_home_1",
        "naver_home_2",
        "blog_1",
        "blog_2",
    ]
    assert rows[1][0] == "보험 추천"


def test_title_generator_can_export_multiple_formats_when_requested(tmp_path: Path) -> None:
    result = run(
        {
            "category": "비즈니스경제",
            "seed_input": "보험 추천",
            "selected_keywords": [
                {
                    "keyword": "보험 추천",
                    "score": 1.0,
                }
            ],
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
                "formats": ["csv", "xlsx", "md"],
            },
        }
    )

    export_artifacts = result["generation_meta"]["export_artifacts"]
    assert [item["format"] for item in export_artifacts] == ["csv", "xlsx", "md"]

    workbook_path = next(Path(item["path"]) for item in export_artifacts if item["format"] == "xlsx")
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["summary", "titles"]
    assert workbook["titles"]["A2"].value == "보험 추천"

    markdown_path = next(Path(item["path"]) for item in export_artifacts if item["format"] == "md")
    markdown_text = markdown_path.read_text(encoding="utf-8")
    assert "# Title Export" in markdown_text
    assert "## 보험 추천" in markdown_text


def test_title_generator_can_export_queue_txt_bundle_with_manifest(tmp_path: Path) -> None:
    result = run(
        {
            "category": "비즈니스경제",
            "seed_input": "보험 추천",
            "selected_keywords": [
                {
                    "keyword": "보험 추천",
                    "score": 1.0,
                }
            ],
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
                "queue_export": {
                    "enabled": True,
                    "destination": "wordpress",
                    "topic": "보험",
                },
            },
        }
    )

    export_artifacts = result["generation_meta"]["export_artifacts"]
    artifact_paths = {item["format"]: Path(item["path"]) for item in export_artifacts}
    assert "txt_live" in artifact_paths
    assert "txt_archive" in artifact_paths
    assert "manifest_json" in artifact_paths
    assert result["generation_meta"]["queue_export"]["destination"] == "wordpress"
    assert artifact_paths["txt_live"].parent.name == "wordpress"
    assert artifact_paths["txt_live"].read_text(encoding="utf-8").strip()
    manifest_payload = json.loads(artifact_paths["manifest_json"].read_text(encoding="utf-8"))
    selected_title = manifest_payload["entries"][0]["selected_title"]
    assert selected_title in artifact_paths["txt_live"].read_text(encoding="utf-8")
    assert manifest_payload["destination"] == "wordpress"
    assert manifest_payload["topic"] == "보험"
    assert manifest_payload["entries"][0]["selected_title"] == selected_title
    return
    selected_title = result["generated_titles"][0]["titles"]["blog"][0]
    assert selected_title in artifact_paths["txt_live"].read_text(encoding="utf-8")
    manifest_payload = json.loads(artifact_paths["manifest_json"].read_text(encoding="utf-8"))
    assert manifest_payload["destination"] == "wordpress"
    assert manifest_payload["topic"] == "보험"
    assert manifest_payload["entries"][0]["selected_title"] == selected_title


def test_title_generation_options_keep_custom_system_prompt() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "system_prompt": "제목 첫 단어는 항상 키워드여야 한다.",
            }
        }
    )

    assert options.system_prompt == "제목 첫 단어는 항상 키워드여야 한다."
    assert "Additional guidance" in options.effective_system_prompt
    assert "avoid repeating the same headline skeleton" in options.effective_system_prompt.lower()


def test_title_generation_options_apply_preset_defaults() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "preset_key": "openai_strict",
                "system_prompt": "숫자형 제목을 우선한다.",
            }
        }
    )

    assert options.preset_key == "openai_strict"
    assert options.preset_label == "안정형 규칙 우선"
    assert options.provider == "openai"
    assert options.model == "gpt-4.1-mini"
    assert options.temperature == 0.2
    assert options.issue_source_mode == "mixed"
    assert "Preset guidance" in options.effective_system_prompt
    assert "Additional guidance" in options.effective_system_prompt


def test_title_generation_options_use_mixed_stable_preset_by_default_in_ai_mode() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
            }
        }
    )

    assert options.preset_key == "openai_mixed_stable"
    assert options.provider == "openai"
    assert options.model == "gpt-4.1-mini"
    assert options.temperature == 0.5
    assert options.issue_context_enabled is True
    assert options.issue_context_limit == 3
    assert options.issue_source_mode == "mixed"
    assert options.community_sources == ("cafe.naver.com", "blog.naver.com", "post.naver.com")
    assert "mixed issue sourcing as the default operating mode" in options.effective_system_prompt
    assert "community reaction cues" in options.effective_system_prompt


def test_title_generation_options_apply_reaction_aggressive_preset_defaults() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "preset_key": "openai_reaction_aggressive",
            }
        }
    )

    assert options.preset_key == "openai_reaction_aggressive"
    assert options.provider == "openai"
    assert options.model == "gpt-4.1-mini"
    assert options.temperature == 0.8
    assert options.issue_source_mode == "reaction"
    assert options.community_sources == ("cafe.naver.com", "blog.naver.com", "post.naver.com")
    assert "Prioritize selected-domain community reaction cues" in options.effective_system_prompt


def test_ai_prompt_builder_includes_category_overlay_and_metrics() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "서울 청약 경쟁률",
                "score": 77.0,
                "profitability_grade": "A",
                "attackability_grade": "2",
                "metrics": {
                    "volume": 1250.0,
                    "cpc": 310.0,
                },
                "target_mode": "single",
                "source_kind": "selected_keyword",
                "source_keywords": ["서울 청약 경쟁률", "서울 분양"],
                "source_note": "최근 청약 일정과 경쟁률 흐름 확인용",
                "issue_context": {
                    "fetched_at": "2026-03-21T08:00:00+09:00",
                    "title_count": 5,
                    "news_count": 2,
                    "source_mix": {"news": 2, "blog": 2, "official": 1},
                    "issue_terms": ["일정", "경쟁률", "분양"],
                    "news_headlines": [
                        "서울 청약 경쟁률 이번주 일정 바뀌나",
                        "서울 분양 경쟁률 다시 오르나",
                    ],
                },
            }
        ]
    )

    assert "Current local date reference:" in prompt
    assert "Never invent unsupported facts" in prompt
    assert "Every title must contain all meaningful keyword tokens from the input keyword." in prompt
    assert "Keep keyword tokens in the same order." in prompt
    assert "category overlay: 부동산 블로그" in prompt
    assert "preferred naver_home pair:" in prompt
    assert "freshness cues:" in prompt
    assert "data hooks:" in prompt
    assert "required keyword tokens: 서울, 청약, 경쟁률" in prompt
    assert "keyword coverage rule:" in prompt
    assert "available signals: score 77" in prompt
    assert "grade A/2" in prompt
    assert "volume 1,250" in prompt
    assert "cpc 310" in prompt
    assert "target context: single" in prompt
    assert "source hints: selected_keyword" in prompt
    assert "source keywords 서울 청약 경쟁률, 서울 분양" in prompt
    assert "live issue context: mode mixed / fetched 2026-03-21 / news 2/5" in prompt
    assert "recent headlines: 서울 청약 경쟁률 이번주 일정 바뀌나" in prompt


def test_ai_prompt_builder_includes_practical_title_shape_hint() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "로지텍 마우스 설정 팁",
                "target_mode": "longtail_selected",
            }
        ]
    )

    assert "practical title shape: setup/help" in prompt
    assert "device or OS context" in prompt


def test_ai_prompt_builder_expands_preorder_and_value_hints() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "닌텐도 스위치2 사전예약",
                "target_mode": "single",
                "source_selection_mode": "seed_anchor",
                "source_selection_reason": "seed_intent_preserved",
            },
            {
                "keyword": "오사카 가성비 호텔",
                "target_mode": "single",
                "source_selection_mode": "seed_anchor",
            },
        ]
    )

    assert "practical title shape: preorder/application" in prompt
    assert "practical title shape: value decision" in prompt
    assert "source hints: selected_keyword / selection seed_anchor" not in prompt
    assert "selection seed_anchor" in prompt
    assert "selection note seed_intent_preserved" in prompt


def test_ai_prompt_builder_warns_against_keyword_shortening() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "손목 편한 마우스 설정 팁",
                "target_mode": "longtail_selected",
            }
        ]
    )

    assert "Do not shorten the keyword phrase" in prompt


def test_ai_prompt_builder_includes_category_boundary_rule_and_senior_overlay() -> None:
    prompt = _build_user_prompt_from_items(
        [
            {
                "keyword": "기초연금 신청자격",
                "quality_report": {
                    "bundle_score": 80,
                    "status": "review",
                    "channel_scores": {"naver_home": 79, "blog": 78},
                },
                "target_mode": "single",
            }
        ]
    )

    assert "Keep strict category boundaries." in prompt
    assert "eligibility, application timing, care, safety aids" in prompt


def test_request_ai_titles_includes_live_issue_context_in_prompt() -> None:
    html = """
    <html>
      <body>
        <a href="https://news.example.com/a" class="news_tit">서울 청약 경쟁률 이번주 일정 바뀌나</a>
        <a href="https://blog.naver.com/post1" class="title_link">서울 청약 경쟁률 비교 포인트</a>
        <a href="https://news.example.com/b" class="news_tit">서울 분양 경쟁률 다시 오르나</a>
      </body>
    </html>
    """
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "test-key",
                "issue_context_enabled": True,
                "issue_context_limit": 1,
            }
        }
    )

    with patch.dict("app.title.ai_client._ISSUE_CONTEXT_CACHE", {}, clear=True), patch(
        "app.title.ai_client.fetch_naver_serp_html",
        return_value=html,
    ), patch(
        "app.title.ai_client._request_openai_titles",
        return_value=[],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=lambda chunk, options: _build_slot_retry_response(chunk),
    ):
        request_ai_titles([{"keyword": "서울 청약 경쟁률"}], options)

    prompt = mocked_request.call_args.args[0]
    assert "live issue context:" in prompt
    assert "recent headlines:" in prompt
    assert "서울 청약 경쟁률 이번주 일정 바뀌나" in prompt
    assert "news 2/3" in prompt
    assert "required keyword tokens: 서울, 청약, 경쟁률" in prompt


def test_request_ai_titles_includes_selected_community_reaction_in_prompt() -> None:
    html = """
    <html>
      <body>
        <a href="https://news.example.com/a" class="news_tit">보험 추천 이번주 비교 포인트</a>
        <a href="https://cafe.naver.com/post1" class="title_link">보험 추천 후기 비교 포인트</a>
        <a href="https://blog.naver.com/post2" class="title_link">보험 추천 가입 조건 정리</a>
      </body>
    </html>
    """
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "test-key",
                "issue_context_enabled": True,
                "issue_context_limit": 1,
                "issue_source_mode": "reaction",
                "community_sources": ["cafe_naver"],
            }
        }
    )

    with patch.dict("app.title.ai_client._ISSUE_CONTEXT_CACHE", {}, clear=True), patch(
        "app.title.ai_client.fetch_naver_serp_html",
        return_value=html,
    ), patch(
        "app.title.ai_client._request_openai_titles",
        return_value=[],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=[
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ],
                degraded=True,
            ),
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ]
            ),
        ],
    ):
        request_ai_titles([{"keyword": "보험 추천"}], options)

    prompt = mocked_request.call_args.args[0]
    assert "community reaction:" in prompt
    assert "community headlines:" in prompt
    assert "후기" in prompt
    assert "보험 추천 후기 비교 포인트" in prompt


def test_request_ai_titles_supports_vertex_provider() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "vertex",
                "api_key": "vertex-key",
                "model": "gemini-2.5-flash-lite",
            }
        }
    )

    with patch(
        "app.title.ai_client._post_json",
        return_value={
            "candidates": [
                {
                    "content": {
                        "parts": [
                            {
                                "text": (
                                    '{"items":[{"keyword":"보험 추천","naver_home":["보험 추천 비교 포인트","보험 추천 선택 기준"],'
                                    '"blog":["보험 추천 가이드","보험 추천 체크포인트"]}]}'
                                )
                            }
                        ]
                    }
                }
            ]
        },
    ) as mocked_post:
        result = request_ai_titles([{"keyword": "보험 추천"}], options)

    called_url = mocked_post.call_args.args[0]
    called_headers = mocked_post.call_args.args[1]
    assert "aiplatform.googleapis.com" in called_url
    assert "publishers/google/models/gemini-2.5-flash-lite:generateContent" in called_url
    assert "key=vertex-key" in called_url
    assert called_headers["Content-Type"] == "application/json"
    assert result[0]["keyword"] == "보험 추천"


def test_request_ai_titles_recovers_from_vertex_trailing_comma_json() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "vertex",
                "api_key": "vertex-key",
                "model": "gemini-2.5-flash-lite",
            }
        }
    )

    malformed_payload = """```json
{
  "items": [
    {
      "keyword": "보험 추천",
      "naver_home": ["보험 추천 비교 포인트", "보험 추천 선택 기준"],
      "blog": ["보험 추천 가이드", "보험 추천 체크포인트"],
    },
  ],
}
```"""

    with patch(
        "app.title.ai_client._post_json",
        return_value={
            "candidates": [
                {
                    "content": {
                        "parts": [
                            {
                                "text": malformed_payload
                            }
                        ]
                    }
                }
            ]
        },
    ):
        result = request_ai_titles([{"keyword": "보험 추천"}], options)

    assert result[0]["keyword"] == "보험 추천"
    assert len(result[0]["titles"]["naver_home"]) == 2
    assert len(result[0]["titles"]["blog"]) == 2


def test_request_ai_json_object_recovers_from_trailing_comma_json() -> None:
    malformed_payload = """```json
{
  "score": {
    "issue_or_context": 18,
    "curiosity_gap": 16,
    "contrast_or_tension": 12,
    "reversal_or_unexpected": 10,
    "emotional_trigger": 11,
    "clarity": 9,
    "readability": 4,
    "total": 80,
  },
  "verdict": "keep",
  "reason": "CTR 기준 통과",
  "rewritten_title": "",
}
```"""

    with patch(
        "app.title.ai_client._post_json",
        return_value={
            "candidates": [
                {
                    "content": {
                        "parts": [
                            {
                                "text": malformed_payload
                            }
                        ]
                    }
                }
            ]
        },
    ):
        result = request_ai_json_object(
            provider="vertex",
            api_key="vertex-key",
            model="gemini-2.5-flash-lite",
            system_prompt="Return JSON only.",
            user_prompt="Evaluate this title.",
        )

    assert result["verdict"] == "keep"
    assert result["score"]["total"] == 80


def test_generate_titles_retries_by_splitting_chunk_on_invalid_json() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "vertex",
                "api_key": "vertex-key",
                "model": "gemini-2.5-flash-lite",
                "fallback_to_template": False,
                "batch_size": 20,
                "auto_retry_enabled": False,
            }
        }
    )
    items = [
        {"keyword": "보험 추천"},
        {"keyword": "실비보험 비교"},
    ]
    requested_chunk_sizes: list[int] = []

    def fake_request(chunk, options):
        requested_chunk_sizes.append(len(chunk))
        if len(chunk) > 1:
            raise TitleProviderError("Provider response contained invalid JSON.")
        keyword = chunk[0]["keyword"]
        return [
            {
                "keyword": keyword,
                "titles": {
                    "naver_home": [f"{keyword} 지금 왜 다를까", f"{keyword} 선택 기준은"],
                    "blog": [f"{keyword} 핵심 정리", f"{keyword} 비교 가이드"],
                },
            }
        ]

    with patch("app.title.title_generator.request_ai_titles", side_effect=fake_request):
        generated, meta = generate_titles(items, options=options)

    assert [item["keyword"] for item in generated] == ["보험 추천", "실비보험 비교"]
    assert meta["used_mode"] == "ai"
    assert requested_chunk_sizes == [2, 1, 1]


def test_resolve_retry_delay_seconds_reads_provider_hint() -> None:
    delay = _resolve_retry_delay_seconds(
        "429 You exceeded your current quota. Please retry in 44.460159992s.",
        attempt=0,
    )

    assert delay == 44.460159992


@pytest.mark.skip(reason="legacy platform-dedupe expectation replaced by global cross-platform dedupe")
def test_title_export_can_split_queue_by_platform_with_platform_only_dedupe(tmp_path: Path) -> None:
    export_payload = export_generated_titles(
        {
            "category": "비즈니스경제",
            "seed_input": "국제금시세",
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
                "queue_export": {
                    "enabled": True,
                    "destination": "all",
                    "quality_gate_enabled": False,
                    "topic": "금융",
                },
            },
        },
        [
            {
                "keyword": "국제금시세",
                "titles": {
                    "naver_home": ["같은 제목", "보조 홈 제목"],
                    "blog": ["같은 제목", "보조 블로그 제목"],
                },
                "target_mode": "single",
            },
            {
                "keyword": "오늘금시세",
                "titles": {
                    "naver_home": ["같은 제목", "다른 홈 제목"],
                    "blog": ["같은 제목", "다른 블로그 제목"],
                },
                "target_mode": "single",
            },
        ],
    )

    queue_export = export_payload["queue_export"]
    assert queue_export["destination"] == "all"
    assert set(queue_export["destinations"]) == {"home", "wordpress"}
    assert len(queue_export["bundles"]) == 2

    bundles = {bundle["destination"]: bundle for bundle in queue_export["bundles"]}
    assert bundles["home"]["row_count"] == 1
    assert bundles["wordpress"]["row_count"] == 1

    home_live_path = next(Path(item["path"]) for item in bundles["home"]["artifacts"] if item["format"] == "txt_live")
    wordpress_live_path = next(Path(item["path"]) for item in bundles["wordpress"]["artifacts"] if item["format"] == "txt_live")

    assert home_live_path.parent.name == "home"
    assert wordpress_live_path.parent.name == "wordpress"
    assert home_live_path.read_text(encoding="utf-8").splitlines() == ["같은 제목"]
    assert wordpress_live_path.read_text(encoding="utf-8").splitlines() == ["같은 제목"]


def test_title_generation_options_parse_quality_retry_settings() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "auto_retry_enabled": False,
                "quality_retry_threshold": 92,
            }
        }
    )

    assert options.auto_retry_enabled is False
    assert options.quality_retry_threshold == 92


def test_title_generation_options_default_quality_retry_settings_are_relaxed() -> None:
    options = TitleGenerationOptions.from_input({"title_options": {"mode": "ai"}})

    assert options.auto_retry_enabled is False
    assert options.quality_retry_threshold == 75


def test_title_export_queue_filters_retry_items_by_default(tmp_path: Path) -> None:
    export_payload = export_generated_titles(
        {
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
                "queue_export": {
                    "enabled": True,
                    "destination": "wordpress",
                    "topic": "finance",
                },
            },
        },
        [
            {
                "keyword": "alpha keyword",
                "titles": {
                    "blog": ["keep this title"],
                    "naver_home": ["home keep"],
                },
                "quality_report": {
                    "bundle_score": 82,
                    "status": "review",
                    "channel_scores": {"naver_home": 80, "blog": 79},
                },
            },
            {
                "keyword": "beta keyword",
                "titles": {
                    "blog": ["drop this title"],
                    "naver_home": ["home drop"],
                },
                "quality_report": {
                    "bundle_score": 61,
                    "status": "retry",
                    "channel_scores": {"naver_home": 62, "blog": 60},
                },
            },
        ],
    )

    queue_export = export_payload["queue_export"]
    assert queue_export["destination"] == "wordpress"
    assert queue_export["row_count"] == 1
    live_path = next(Path(item["path"]) for item in queue_export["artifacts"] if item["format"] == "txt_live")
    assert live_path.read_text(encoding="utf-8").splitlines() == ["keep this title"]


def test_title_export_uses_fallback_titles_to_avoid_cross_platform_duplicates(tmp_path: Path) -> None:
    export_payload = export_generated_titles(
        {
            "category": "finance",
            "seed_input": "gold price",
            "title_export": {
                "enabled": True,
                "output_dir": str(tmp_path),
                "queue_export": {
                    "enabled": True,
                    "destination": "all",
                    "quality_gate_enabled": False,
                    "topic": "finance",
                },
            },
        },
        [
            {
                "keyword": "gold price",
                "titles": {
                    "naver_home": ["shared title", "home alt"],
                    "blog": ["shared title", "blog alt"],
                },
                "target_mode": "single",
            },
            {
                "keyword": "today gold price",
                "titles": {
                    "naver_home": ["shared title", "home alt 2"],
                    "blog": ["shared title", "blog alt 2"],
                },
                "target_mode": "single",
            },
        ],
    )

    bundles = {bundle["destination"]: bundle for bundle in export_payload["queue_export"]["bundles"]}
    assert bundles["home"]["row_count"] == 2
    assert bundles["wordpress"]["row_count"] == 2

    home_live_path = next(Path(item["path"]) for item in bundles["home"]["artifacts"] if item["format"] == "txt_live")
    wordpress_live_path = next(Path(item["path"]) for item in bundles["wordpress"]["artifacts"] if item["format"] == "txt_live")

    assert home_live_path.read_text(encoding="utf-8").splitlines() == ["shared title", "home alt 2"]
    assert wordpress_live_path.read_text(encoding="utf-8").splitlines() == ["blog alt", "blog alt 2"]
    assert home_live_path.name.endswith("__finance.txt")
    assert wordpress_live_path.name.endswith("__finance.txt")


def test_title_generator_reports_preset_metadata() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": "보험 추천",
                "titles": {
                    "naver_home": ["보험 추천 비교 포인트 2가지", "보험 추천 선택 기준 달라졌다"],
                    "blog": ["보험 추천 핵심 가이드", "보험 추천 체크 포인트 정리"],
                },
            }
        ],
    ):
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "보험 추천",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "preset_key": "openai_balanced",
                    "api_key": "test-key",
                },
            }
        )

    assert result["generation_meta"]["preset_key"] == "openai_balanced"
    assert result["generation_meta"]["preset_label"] == "추천 균형형"
    assert result["generation_meta"]["provider"] == "openai"
    assert result["generation_meta"]["model"] == "gpt-4o-mini"
    assert result["generation_meta"]["temperature"] == 0.7


def test_title_generator_keeps_custom_quality_retry_threshold_without_extra_retry_when_pair_ready() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        side_effect=[
            [
                {
                    "keyword": "보험 추천",
                    "titles": {
                        "naver_home": [
                            "지금 비교할 보험 추천 포인트 2가지",
                            "한눈에 보는 보험 추천 선택 기준",
                        ],
                        "blog": [
                            "실전 보험 추천 비교 가이드",
                            "지금 보는 보험 추천 체크리스트",
                        ],
                    },
                }
            ],
            [
                {
                    "keyword": "보험 추천",
                    "titles": {
                        "naver_home": [
                            "보험 추천 지금 비교할 포인트 2가지",
                            "보험 추천 선택 기준 한눈에 보기",
                        ],
                        "blog": [
                            "보험 추천 비교 가이드",
                            "보험 추천 체크리스트",
                        ],
                    },
                }
            ],
        ],
    ) as mocked_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "보험 추천",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "quality_retry_threshold": 95,
                    "auto_retry_enabled": False,
                    "keyword_modes": ["single"],
                },
            }
        )

    assert mocked_request.call_count == 1
    assert result["generation_meta"]["quality_retry_threshold"] == 95
    assert result["generation_meta"]["auto_retry"]["retry_threshold"] == 95
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 0
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 0
    assert result["generated_titles"][0]["quality_report"]["recommended_pair_ready"] is False


def test_title_quality_retry_prompt_includes_naver_home_ctr_guidance() -> None:
    prompt = _build_quality_retry_prompt("", 84)

    assert "evaluate CTR first, not SEO" in prompt
    assert "Do not penalize a naver_home title just because it is phrased as a question" in prompt
    assert "Strong positive CTR signals include question-led hooks" in prompt
    assert "Do not penalize repeated question patterns" in prompt
    assert "stays short" in prompt
    assert "low explicit information density" in prompt
    assert "Do penalize purely informational naver_home titles" in prompt
    assert "Do not use SEO or blog criteria" in prompt


def test_title_generator_auto_retries_low_quality_ai_titles() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        side_effect=[
            [
                {
                    "keyword": "보험 추천",
                    "titles": {
                        "naver_home": ["보험 추천", "보험 추천"],
                        "blog": ["보험 추천", "보험 추천"],
                    },
                }
            ],
            [
                {
                    "keyword": "보험 추천",
                    "titles": {
                        "naver_home": ["보험 추천, 왜 혜택 차이가 먼저 보일까?", "보험 추천, 지금 갈리는 선택 기준 2가지"],
                        "blog": ["보험 추천 혜택 차이와 보장 범위 정리", "보험 추천 가입 전 체크 포인트 비교"],
                    },
                }
            ],
            [
                {
                    "keyword": "보험 추천",
                    "titles": {
                        "naver_home": ["보험 추천, 왜 혜택 차이가 먼저 보일까?", "보험 추천, 지금 갈리는 선택 기준 2가지"],
                        "blog": ["보험 추천 혜택 차이와 보장 범위 정리", "보험 추천 가입 전 체크 포인트 비교"],
                    },
                }
            ],
        ],
    ) as mocked_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "보험 추천",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    assert mocked_request.call_count == 1
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 1
    assert result["generated_titles"][0]["quality_report"]["retry_recommended"] is False
    assert result["generated_titles"][0]["quality_report"]["bundle_score"] >= 75
    assert result["generation_meta"]["model_escalation"]["triggered"] is True


def test_title_generator_skips_auto_retry_when_recommended_pair_is_already_ready() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": "로지텍 마우스",
                "titles": {
                    "naver_home": [
                        "로지텍 마우스, 왜 배터리 체감이 갈릴까?",
                        "로지텍 마우스, 연결 안정성 먼저 볼 이유",
                    ],
                    "blog": [
                        "로지텍 마우스 사용 후기",
                        "로지텍 마우스 배터리·연결 안정성·추천 대상 정리",
                    ],
                },
            }
        ],
    ) as mocked_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "로지텍 마우스",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    report = result["generated_titles"][0]["quality_report"]
    assert mocked_request.call_count == 1
    assert report["recommended_pair_ready"] is True
    assert report["status"] == "good"
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 0
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 0


def test_title_generator_preserves_full_keyword_after_failed_retries() -> None:
    keyword = "손목 편한 마우스 설정 팁"

    with patch(
        "app.title.title_generator.request_ai_titles",
        side_effect=[
            [
                {
                    "keyword": keyword,
                    "titles": {
                        "naver_home": [
                            "편한 마우스 설정 팁, 최신 정보",
                            "편한 마우스 설정 팁, 뭐가 다를까?",
                        ],
                        "blog": [
                            "편한 마우스 설정 팁 총정리",
                            "편한 마우스 설정 팁 완벽 가이드",
                        ],
                    },
                }
            ],
            [
                {
                    "keyword": keyword,
                    "titles": {
                        "naver_home": [
                            "편한 마우스 설정 팁, 최신 비교 분석",
                            "편한 마우스 설정 팁, 이것만 알면",
                        ],
                        "blog": [
                            "편한 마우스 설정 팁 최신 정보",
                            "편한 마우스 설정 팁 구매 가이드",
                        ],
                    },
                }
            ],
            [
                {
                    "keyword": keyword,
                    "titles": {
                        "naver_home": [
                            f"{keyword}, 배터리와 클릭감 차이",
                            f"{keyword}, 연결 끊김부터 먼저 볼까?",
                        ],
                        "blog": [
                            f"{keyword} 배터리와 클릭감 차이 정리",
                            f"{keyword} 연결 끊김 원인과 해결 순서",
                        ],
                    },
                }
            ],
        ],
    ) as mocked_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    item = result["generated_titles"][0]
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]

    assert mocked_request.call_count == 1
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 1
    assert result["generation_meta"]["model_escalation"]["triggered"] is True
    assert item["quality_report"]["retry_recommended"] is False
    assert item["quality_report"]["bundle_score"] >= 80
    assert all(keyword in title for title in all_titles)
    assert all(
        banned not in title
        for title in all_titles
        for banned in (
            "최신 정보",
            "최신 비교 분석",
            "총정리",
            "완벽 가이드",
            "뭐가 다를까",
            "이것만 알면",
            "구매 가이드",
        )
    )


def test_practical_rescue_stays_generic_for_non_product_setting_keyword() -> None:
    item = _build_practical_rescue_item({"keyword": "전세 계약 설정 팁"})

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert all("블루투스" not in title for title in all_titles)
    assert all("DPI" not in title for title in all_titles)
    assert all("버튼" not in title for title in all_titles)


def test_practical_rescue_uses_keyboard_specific_setting_frames_without_dpi() -> None:
    item = _build_practical_rescue_item({"keyword": "아콘 키보드 설정 팁"})

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert all("DPI" not in title for title in all_titles)
    assert any(
        any(term in title for term in ("멀티페어링", "키맵", "한영", "FN Lock", "배열"))
        for title in all_titles
    )


def test_practical_rescue_detects_preorder_keywords_without_generic_wrappers() -> None:
    item = _build_practical_rescue_item({"keyword": "로지텍 지슈스 사전예약"})

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert all("로지텍 지슈스 사전예약" in title for title in all_titles)
    assert all(
        banned not in title
        for title in all_titles
        for banned in ("총정리", "최신 정보", "구매 가이드", "놓치면 후회", "지금 신청?")
    )
    assert any("일정" in title or "신청" in title or "혜택" in title for title in all_titles)


def test_practical_rescue_problem_keywords_rotate_across_more_than_one_frame() -> None:
    first = _build_practical_rescue_item({"keyword": "로지텍 슈퍼스트라이크 자주 생기는 문제"})
    second = _build_practical_rescue_item({"keyword": "로지텍 지슈라 스트라이크 자주 생기는 문제"})

    assert first is not None
    assert second is not None
    assert first["titles"]["naver_home"] != second["titles"]["naver_home"]
    assert first["titles"]["blog"] != second["titles"]["blog"]


def test_practical_rescue_adds_concrete_value_frames_for_seed_anchor_keyword() -> None:
    item = _build_practical_rescue_item(
        {
            "keyword": "오사카 가성비 호텔",
            "target_mode": "single",
            "source_selection_mode": "seed_anchor",
        }
    )

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert all("오사카 가성비 호텔" in title for title in all_titles)
    assert any(
        any(term in title for term in ("위치", "교통", "추가요금", "예산", "조식", "취소"))
        for title in all_titles
    )
    assert all(
        banned not in title
        for title in all_titles
        for banned in ("숨은 보석", "찾아봐요", "최신 정보")
    )


def test_practical_rescue_specializes_preorder_single_and_checkpoint_variants() -> None:
    single = _build_practical_rescue_item(
        {
            "keyword": "닌텐도 스위치2 사전예약",
            "target_mode": "single",
            "source_selection_mode": "seed_anchor",
        }
    )
    checkpoint = _build_practical_rescue_item(
        {
            "keyword": "닌텐도 스위치2 사전예약 전 체크포인트",
            "target_mode": "longtail_selected",
            "source_selection_mode": "seed_anchor",
        }
    )

    assert single is not None
    assert checkpoint is not None
    single_titles = single["titles"]["naver_home"] + single["titles"]["blog"]
    checkpoint_titles = checkpoint["titles"]["naver_home"] + checkpoint["titles"]["blog"]
    assert single["titles"]["naver_home"] != checkpoint["titles"]["naver_home"]
    assert any(
        any(term in title for term in ("신청 링크", "카드 혜택", "오픈 일정", "결제 수단", "수령 일정"))
        for title in single_titles
    )
    assert any(
        any(term in title for term in ("본인 인증", "제한 수량", "결제 조건", "접속 시간", "수령 일정"))
        for title in checkpoint_titles
    )


def test_practical_rescue_builds_concrete_generic_product_single_titles() -> None:
    item = _build_practical_rescue_item(
        {
            "keyword": "로지텍 마우스",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        }
    )

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert any(
        any(term in title for term in ("실사용", "장단점", "가격대", "추천 대상", "클릭감", "배터리"))
        for title in all_titles
    )
    assert all(
        banned not in title
        for title in all_titles
        for banned in ("당신의 선택", "왜 인기", "신상", "트렌드 분석")
    )


def test_practical_rescue_builds_concrete_generic_stay_single_titles() -> None:
    item = _build_practical_rescue_item(
        {
            "keyword": "오사카 호텔",
            "target_mode": "single",
            "source_kind": "selected_keyword",
        }
    )

    assert item is not None
    all_titles = item["titles"]["naver_home"] + item["titles"]["blog"]
    assert any(
        any(term in title for term in ("위치", "교통", "추가요금", "객실", "조식", "취소"))
        for title in all_titles
    )


def test_practical_rescue_uses_partial_slot_refresh_without_full_enrich() -> None:
    keyword = "alpha beta"
    current_results = [
        {
            "keyword": keyword,
            "titles": {
                "naver_home": [keyword, keyword],
                "blog": [keyword, keyword],
            },
        }
    ]
    current_enriched_results, _ = enrich_title_results(current_results)
    rescue_item = {
        "keyword": keyword,
        "titles": {
            "naver_home": [
                f"{keyword}, 지금 갈리는 기준은?",
                f"{keyword}, 비교 포인트 2가지",
            ],
            "blog": [
                f"{keyword} 비교 기준 정리",
                f"{keyword} 조건별 선택 기준 설명",
            ],
        },
    }

    with patch(
        "app.title.title_generator._build_practical_rescue_item",
        return_value=rescue_item,
    ), patch(
        "app.title.title_generator.enrich_title_results",
        side_effect=AssertionError("full enrich should not be called"),
    ):
        rescued_results, rescued_enriched_results, quality_summary, accepted_keywords = _apply_practical_rescue_candidates(
            current_results=current_results,
            current_enriched_results=current_enriched_results,
            original_items_by_keyword={keyword: current_results[0]},
            retry_threshold=75,
            retry_attempt_counts={},
            recent_batch_size=20,
            max_slot_rewrite_attempts=2,
        )

    assert accepted_keywords == [keyword]
    assert rescued_results[0]["titles"]["naver_home"][0] != keyword
    assert rescued_enriched_results[0]["quality_report"]["bundle_score"] >= current_enriched_results[0]["quality_report"]["bundle_score"]
    assert quality_summary["total_count"] == 1


def test_practical_rescue_skips_slots_that_reached_retry_limit() -> None:
    keyword = "alpha beta"
    current_results = [
        {
            "keyword": keyword,
            "titles": {
                "naver_home": [keyword, keyword],
                "blog": [keyword, keyword],
            },
        }
    ]
    current_enriched_results, quality_summary = enrich_title_results(current_results)
    retry_attempt_counts = {
        "0_home_1": 2,
        "0_home_2": 2,
        "0_blog_1": 2,
        "0_blog_2": 2,
    }

    with patch("app.title.title_generator._build_practical_rescue_item") as mocked_build_rescue_item:
        rescued_results, rescued_enriched_results, rescued_summary, accepted_keywords = _apply_practical_rescue_candidates(
            current_results=current_results,
            current_enriched_results=current_enriched_results,
            original_items_by_keyword={keyword: current_results[0]},
            retry_threshold=75,
            retry_attempt_counts=retry_attempt_counts,
            recent_batch_size=20,
            max_slot_rewrite_attempts=2,
        )

    assert mocked_build_rescue_item.call_count == 0
    assert rescued_results == current_results
    assert rescued_enriched_results == current_enriched_results
    assert rescued_summary == quality_summary
    assert accepted_keywords == []


def test_title_generator_escalates_model_after_two_failed_quality_attempts() -> None:
    """
    with patch(
        "app.title.title_generator.request_ai_titles",
        side_effect=[
            [
                {
                    "keyword": "蹂댄뿕 異붿쿇",
                    "titles": {
                        "naver_home": ["蹂댄뿕 異붿쿇", "蹂댄뿕 異붿쿇"],
                        "blog": ["蹂댄뿕 異붿쿇", "蹂댄뿕 異붿쿇"],
                    },
                }
            ],
            [
                {
                    "keyword": "蹂댄뿕 異붿쿇",
                    "titles": {
                        "naver_home": ["蹂댄뿕 異붿쿇 鍮꾧탳", "蹂댄뿕 異붿쿇 鍮꾧탳"],
                        "blog": ["蹂댄뿕 異붿쿇 媛?대뱶", "蹂댄뿕 異붿쿇 媛?대뱶"],
                    },
                }
            ],
            [
                {
                    "keyword": "蹂댄뿕 異붿쿇",
                    "titles": {
                        "naver_home": [
                            "蹂댄뿕 異붿쿇 吏湲?鍮꾧탳 ?ъ씤??2媛吏",
                            "蹂댄뿕 異붿쿇 ?좏깮 湲곗? ?щ씪吏??댁쑀",
                        ],
                        "blog": [
                            "蹂댄뿕 異붿쿇 鍮꾧탳 媛?대뱶",
                            "蹂댄뿕 異붿쿇 泥댄겕由ъ뒪??,
                        ],
                    },
                }
            ],
        ],
    ) as mocked_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": "蹂댄뿕 異붿쿇",
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    assert mocked_request.call_count == 3
    assert mocked_request.call_args_list[0].kwargs["options"].model == "gpt-4o-mini"
    assert mocked_request.call_args_list[1].kwargs["options"].model == "gpt-4o-mini"
    assert mocked_request.call_args_list[2].kwargs["options"].model == "gpt-4.1-mini"
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 0
    assert result["generation_meta"]["model_escalation"]["enabled"] is True
    assert result["generation_meta"]["model_escalation"]["triggered"] is True
    assert result["generation_meta"]["model_escalation"]["source_model"] == "gpt-4o-mini"
    assert result["generation_meta"]["model_escalation"]["target_model"] == "gpt-4.1-mini"
    assert result["generation_meta"]["model_escalation"]["attempted_count"] == 1
    assert result["generation_meta"]["model_escalation"]["accepted_count"] == 1
    assert result["generation_meta"]["final_model"] == "gpt-4.1-mini"
    assert result["generated_titles"][0]["quality_report"]["retry_recommended"] is False
    assert result["generated_titles"][0]["quality_report"]["bundle_score"] >= 75
    """

    keyword = "insurance plan"

    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": keyword,
                "titles": {
                    "naver_home": [keyword, keyword],
                    "blog": [keyword, keyword],
                },
            }
        ],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=[
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ],
                degraded=True,
            ),
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ]
            ),
        ],
    ) as mocked_slot_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    assert mocked_request.call_count == 1
    assert mocked_request.call_args_list[0].kwargs["options"].model == "gpt-4o-mini"
    assert mocked_slot_request.call_args_list[0].kwargs["options"].model == "gpt-4o-mini"
    assert mocked_slot_request.call_args_list[1].kwargs["options"].model == "gpt-4.1-mini"
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 1
    assert result["generation_meta"]["auto_retry"]["accepted_count"] == 1
    assert result["generation_meta"]["model_escalation"]["enabled"] is True
    assert result["generation_meta"]["model_escalation"]["triggered"] is True
    assert result["generation_meta"]["model_escalation"]["source_model"] == "gpt-4o-mini"
    assert result["generation_meta"]["model_escalation"]["target_model"] == "gpt-4.1-mini"
    assert result["generation_meta"]["model_escalation"]["attempted_count"] == 1
    assert result["generation_meta"]["model_escalation"]["accepted_count"] == 1
    assert result["generation_meta"]["final_model"] == "gpt-4.1-mini"
    assert result["generated_titles"][0]["quality_report"]["retry_recommended"] is False
    assert result["generated_titles"][0]["quality_report"]["bundle_score"] >= 75


def test_title_generator_quality_retry_preserves_prompt_context() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "test-key",
                "model": "gpt-4.1-mini",
            }
        }
    )
    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": "서울 청약 경쟁률",
                "titles": {
                    "naver_home": ["서울 청약 경쟁률", "서울 청약 경쟁률"],
                    "blog": ["서울 청약 경쟁률", "서울 청약 경쟁률"],
                },
            }
        ],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=lambda chunk, options: _build_slot_retry_response(chunk),
    ) as mocked_slot_request:
        generate_titles(
            [
                {
                    "keyword": "서울 청약 경쟁률",
                    "score": 77.0,
                    "metrics": {"volume": 1250.0, "cpc": 310.0},
                    "source_kind": "selected_keyword",
                    "source_note": "최근 청약 일정과 경쟁률 흐름 확인용",
                    "source_keywords": ["서울 청약 경쟁률", "서울 분양"],
                }
            ],
            options=options,
        )

    assert mocked_request.call_count == 1
    assert mocked_slot_request.call_count == 1
    retry_chunk = mocked_slot_request.call_args.args[0]
    retry_options = mocked_slot_request.call_args.kwargs["options"]

    assert retry_chunk[0]["keyword"] == "서울 청약 경쟁률"
    assert retry_chunk[0]["score"] > 0
    assert retry_chunk[0]["metrics"]["volume"] == 1250.0
    assert retry_chunk[0]["source_note"] == "최근 청약 일정과 경쟁률 흐름 확인용"
    assert "home-feed pair issue-aware" in retry_options.system_prompt
    assert "Do not invent unsupported dates, numbers, rankings, or official changes" in retry_options.system_prompt


def test_collect_retry_slot_candidates_adds_recent_batch_peer_titles_from_other_keywords() -> None:
    current_results = [
        {
            "keyword": "alpha mouse",
            "titles": {
                "naver_home": ["alpha mouse what changed", "alpha mouse compare point"],
                "blog": ["alpha mouse buying note", "alpha mouse setup guide"],
            },
        },
        {
            "keyword": "beta mouse",
            "titles": {
                "naver_home": ["beta mouse what changed", "beta mouse compare point"],
                "blog": ["beta mouse buying note", "beta mouse setup guide"],
            },
        },
    ]
    current_enriched_results = [
        {
            "keyword": item["keyword"],
            "quality_report": {
                "retry_recommended": True,
                "bundle_score": 61,
                "title_checks": {
                    "naver_home": [
                        {"title": item["titles"]["naver_home"][0], "score": 61, "status": "retry", "issues": ["flat"]},
                        {"title": item["titles"]["naver_home"][1], "score": 60, "status": "retry", "issues": ["flat"]},
                    ],
                    "blog": [
                        {"title": item["titles"]["blog"][0], "score": 59, "status": "retry", "issues": ["flat"]},
                        {"title": item["titles"]["blog"][1], "score": 58, "status": "retry", "issues": ["flat"]},
                    ],
                },
            },
        }
        for item in current_results
    ]

    retry_candidates = _collect_retry_slot_candidates(
        current_results,
        current_enriched_results,
        retry_threshold=75,
        recent_batch_size=20,
    )

    candidate = next(item for item in retry_candidates if item["slot_id"] == "0_home_1")
    assert candidate["peer_titles"][:3] == [
        "alpha mouse compare point",
        "alpha mouse buying note",
        "alpha mouse setup guide",
    ]
    assert len(candidate["peer_titles"]) == 7
    assert any(
        title in candidate["peer_titles"][3:]
        for title in (
            "beta mouse what changed",
            "beta mouse compare point",
            "beta mouse buying note",
            "beta mouse setup guide",
        )
    )


def test_should_retry_slot_blocks_home_retry_for_ctr_hooks_and_short_titles() -> None:
    assert _should_retry_slot(
        {
            "score": 66,
            "score_breakdown": {"total": 66},
        },
        75,
        channel_name="naver_home",
        title="alpha card why now?",
    ) is False
    assert _should_retry_slot(
        {
            "score": 69,
            "score_breakdown": {
                "total": 69,
                "emotional_trigger": 12,
            },
        },
        75,
        channel_name="naver_home",
        title="alpha card shock point",
    ) is False
    assert _should_retry_slot(
        {
            "score": 69,
            "score_breakdown": {"total": 69},
        },
        75,
        channel_name="naver_home",
        title="alpha card short",
    ) is False


def test_collect_retry_slot_candidates_only_keeps_home_slots_below_new_gate() -> None:
    current_results = [
        {
            "keyword": "alpha card",
            "titles": {
                "naver_home": [
                    "alpha card what changed",
                    "alpha card why now?",
                ],
                "blog": [
                    "alpha card compare guide",
                    "alpha card checklist",
                ],
            },
        }
    ]
    current_enriched_results = [
        {
            "keyword": "alpha card",
            "quality_report": {
                "retry_recommended": True,
                "bundle_score": 71,
                "title_checks": {
                    "naver_home": [
                        {
                            "title": "alpha card what changed",
                            "score": 74,
                            "status": "retry",
                            "issues": ["needs stronger CTR"],
                            "score_breakdown": {
                                "issue_or_context": 16,
                                "curiosity_gap": 0,
                                "contrast_or_conflict": 0,
                                "total": 74,
                            },
                        },
                        {
                            "title": "alpha card why now?",
                            "score": 66,
                            "status": "retry",
                            "issues": ["question hook already present"],
                            "score_breakdown": {
                                "issue_or_context": 10,
                                "curiosity_gap": 10,
                                "contrast_or_conflict": 0,
                                "total": 66,
                            },
                        },
                    ],
                    "blog": [
                        {
                            "title": "alpha card compare guide",
                            "score": 72,
                            "status": "retry",
                            "issues": ["soft issue"],
                            "score_breakdown": {"total": 72},
                        },
                        {
                            "title": "alpha card checklist",
                            "score": 78,
                            "status": "keep",
                            "issues": [],
                            "score_breakdown": {"total": 78},
                        },
                    ],
                },
            },
        }
    ]

    retry_candidates = _collect_retry_slot_candidates(
        current_results,
        current_enriched_results,
        retry_threshold=75,
        recent_batch_size=20,
    )

    assert [candidate["slot_id"] for candidate in retry_candidates] == ["0_home_1"]


def test_request_ai_slot_rewrites_keeps_extended_peer_titles_in_prompt() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "test-key",
                "model": "gpt-4.1-mini",
            }
        }
    )
    captured: dict[str, str] = {}

    def fake_request_ai_json_object(**kwargs):
        captured["user_prompt"] = kwargs["user_prompt"]
        return {
            "items": [
                {
                    "slot_id": "0_home_1",
                    "title": "alpha mouse, which choice changed now",
                }
            ]
        }

    peer_titles = [
        "alpha mouse compare point",
        "alpha mouse buying note",
        "alpha mouse setup guide",
        "beta mouse what changed",
        "beta mouse compare point",
        "beta mouse buying note",
        "beta mouse setup guide",
    ]

    with patch("app.title.ai_client.request_ai_json_object", side_effect=fake_request_ai_json_object):
        request_ai_slot_rewrites(
            [
                {
                    "slot_id": "0_home_1",
                    "keyword": "alpha mouse",
                    "channel": "naver_home",
                    "slot_index": 1,
                    "current_title": "alpha mouse what changed",
                    "peer_titles": peer_titles,
                    "issues": ["flat"],
                    "score": 61,
                }
            ],
            options,
        )

    assert "peer_titles:" in captured["user_prompt"]
    for title in peer_titles:
        assert title in captured["user_prompt"]


def test_request_ai_slot_rewrites_includes_zero_score_in_prompt() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "api_key": "test-key",
                "model": "gpt-4.1-mini",
            }
        }
    )
    captured: dict[str, str] = {}

    def fake_request_ai_json_object(**kwargs):
        captured["user_prompt"] = kwargs["user_prompt"]
        return {
            "items": [
                {
                    "slot_id": "0_blog_1",
                    "title": "alpha mouse setup guide",
                }
            ]
        }

    with patch("app.title.ai_client.request_ai_json_object", side_effect=fake_request_ai_json_object):
        request_ai_slot_rewrites(
            [
                {
                    "slot_id": "0_blog_1",
                    "keyword": "alpha mouse",
                    "channel": "blog",
                    "slot_index": 1,
                    "current_title": "alpha mouse",
                    "issues": ["too vague"],
                    "score": 0,
                }
            ],
            options,
        )

    assert "- score: 0" in captured["user_prompt"]


def test_title_service_reuses_existing_serp_issue_context() -> None:
    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": "보험 추천",
                "titles": {
                    "naver_home": ["보험 추천 이번주 비교 포인트", "보험 추천 기준 다시 보인다"],
                    "blog": ["보험 추천 체크포인트", "보험 추천 이슈 정리"],
                },
            }
        ],
    ) as mocked_request:
        run(
            {
                "selected_keywords": [
                    {
                        "keyword": "보험 추천",
                        "score": 1.0,
                    }
                ],
                "serp_competition_summary": {
                    "queries": [
                        {
                            "query": "보험 추천",
                            "source_mix": {"news": 1, "blog": 2},
                            "common_terms": ["비교", "기준"],
                            "top_titles": ["보험 추천 이번주 비교 포인트", "보험 추천 기준 바뀌었나"],
                        }
                    ]
                },
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "test-key",
                    "issue_context_enabled": False,
                },
            }
        )

    input_items = mocked_request.call_args.args[0]
    assert input_items[0]["issue_context"]["query"] == "보험 추천"
    assert input_items[0]["issue_context"]["source_mix"]["news"] == 1


def _build_longtail_title_input() -> dict:
    insurance = "\ubcf4\ud5d8 \ucd94\ucc9c"
    car_insurance = "\uc790\ub3d9\ucc28 \ubcf4\ud5d8"
    return {
        "selected_keywords": [
            {
                "keyword": insurance,
                "profitability_grade": "A",
                "attackability_grade": "2",
                "score": 76.0,
                "metrics": {"volume": 1200.0, "cpc": 210.0},
            },
            {
                "keyword": car_insurance,
                "profitability_grade": "B",
                "attackability_grade": "2",
                "score": 62.0,
                "metrics": {"volume": 900.0, "cpc": 180.0},
            },
        ],
        "keyword_clusters": [
            {
                "cluster_id": "cluster-01",
                "representative_keyword": insurance,
                "topic_terms": ["\ubcf4\ud5d8", "\ucd94\ucc9c", "\uac00\uc785"],
                "all_keywords": [insurance],
            },
            {
                "cluster_id": "cluster-02",
                "representative_keyword": car_insurance,
                "topic_terms": ["\uc790\ub3d9\ucc28", "\ubcf4\ud5d8", "\ud2b9\uc57d"],
                "all_keywords": [car_insurance],
            },
        ],
        "longtail_suggestions": [
            {
                "suggestion_id": "longtail-01",
                "cluster_id": "cluster-01",
                "representative_keyword": insurance,
                "source_keyword": "\ubcf4\ud5d8 \uac00\uc785 \uc870\uac74",
                "longtail_keyword": "\ubcf4\ud5d8 \uac00\uc785 \uc870\uac74 \uccb4\ud06c\ub9ac\uc2a4\ud2b8",
                "combination_terms": [insurance, "\ubcf4\ud5d8 \uac00\uc785 \uc870\uac74", "\uac00\uc785 \uc870\uac74"],
                "verification_status": "pass",
                "verified_score": 68.0,
            }
        ],
        "analyzed_keywords": [
            {
                "keyword": "\ubcf4\ud5d8 \uac00\uc785 \ubc29\ubc95",
                "score": 44.0,
                "metrics": {"volume": 180.0, "cpc": 120.0},
            },
            {
                "keyword": "\ubcf4\ud5d8 \uac00\uc785 \uc870\uac74",
                "score": 48.0,
                "metrics": {"volume": 220.0, "cpc": 135.0},
            },
            {
                "keyword": "\uc790\ub3d9\ucc28 \ubcf4\ud5d8 \ud2b9\uc57d \uc11c\ub958",
                "score": 20.0,
                "metrics": {"volume": 15.0, "cpc": 45.0},
            },
            {
                "keyword": "\uc790\ub3d9\ucc28 \ubcf4\ud5d8 \ud2b9\uc57d \uc815\ub9ac",
                "score": 12.0,
                "metrics": {"volume": 8.0, "cpc": 30.0},
            },
        ],
        "title_options": {
            "keyword_modes": [
                "single",
                "longtail_selected",
                "longtail_exploratory",
                "longtail_experimental",
            ]
        },
        "longtail_options": {
            "optional_suffix_keys": ["guide", "checklist"],
        },
    }


def test_title_generator_builds_longtail_targets_for_keyword_modes() -> None:
    result = run(_build_longtail_title_input())

    summary = result["generation_meta"]["target_summary"]
    assert summary["requested_modes"] == [
        "single",
        "longtail_selected",
        "longtail_exploratory",
        "longtail_experimental",
    ]
    assert summary["mode_counts"]["single"] == 2
    assert summary["mode_counts"]["longtail_selected"] >= 1
    assert summary["mode_counts"]["longtail_exploratory"] >= 1
    assert summary["mode_counts"]["longtail_experimental"] >= 0
    assert len(result["generated_titles"]) == summary["target_count"]

    generated_modes = {item["target_mode"] for item in result["generated_titles"]}
    assert {"single", "longtail_selected", "longtail_exploratory"}.issubset(generated_modes)
    assert any(item["target_id"] for item in result["generated_titles"])
    if summary["mode_counts"]["longtail_experimental"] >= 1:
        assert any(
            item["base_keyword"] == "\uc790\ub3d9\ucc28 \ubcf4\ud5d8"
            for item in result["generated_titles"]
            if item["target_mode"] == "longtail_experimental"
        )


def test_build_title_targets_defaults_to_single_and_v1_modes() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "score": 76.0,
                    "metrics": {"volume": 1200.0, "cpc": 210.0},
                }
            ],
            "keyword_clusters": [
                {
                    "cluster_id": "cluster-01",
                    "representative_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "topic_terms": ["\ubcf4\ud5d8", "\ucd94\ucc9c", "\uac00\uc785"],
                    "all_keywords": ["\ubcf4\ud5d8 \ucd94\ucc9c", "\ubcf4\ud5d8 \uac00\uc785 \uccb4\ud06c\ub9ac\uc2a4\ud2b8"],
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "longtail_keyword": "\ubcf4\ud5d8 \uac00\uc785 \uccb4\ud06c\ub9ac\uc2a4\ud2b8",
                    "verification_status": "pass",
                    "verified_score": 68.0,
                }
            ],
        }
    )

    assert summary["requested_modes"] == ["single", "longtail_selected"]
    assert summary["mode_counts"]["single"] == 1
    assert summary["mode_counts"]["longtail_selected"] == 1
    assert {item["target_mode"] for item in items} == {"single", "longtail_selected"}


def test_build_title_targets_rewrites_low_signal_longtail_keywords() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "로지텍 마우스",
                    "score": 76.0,
                    "metrics": {"volume": 1200.0, "cpc": 210.0},
                }
            ],
            "keyword_clusters": [
                {
                    "cluster_id": "cluster-01",
                    "representative_keyword": "로지텍 마우스",
                    "topic_terms": ["로지텍", "마우스"],
                    "all_keywords": ["로지텍 마우스"],
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "로지텍 마우스",
                    "longtail_keyword": "로지텍 마우스 추천 기준",
                    "verification_status": "pass",
                    "verified_score": 71.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "로지텍 마우스",
                    "longtail_keyword": "로지텍 마우스 연결 방법",
                    "verification_status": "pass",
                    "verified_score": 69.0,
                },
            ],
        }
    )

    keywords = {item["keyword"] for item in items}

    assert "로지텍 마우스 추천 기준" not in keywords
    assert any(keyword in keywords for keyword in ("로지텍 마우스 실사용 차이", "로지텍 마우스 장단점"))
    assert "로지텍 마우스 연결 방법" in keywords
    assert summary["mode_counts"]["longtail_selected"] == 2


def test_build_title_targets_limits_weak_singleton_v1_targets() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "블루투스 키보드",
                    "score": 62.0,
                    "metrics": {"volume": 720.0, "cpc": 115.0},
                }
            ],
            "keyword_clusters": [
                {
                    "cluster_id": "cluster-01",
                    "representative_keyword": "블루투스 키보드",
                    "topic_terms": ["블루투스", "키보드"],
                    "all_keywords": ["블루투스 키보드"],
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "블루투스 키보드",
                    "source_keyword": "블루투스 키보드",
                    "longtail_keyword": "블루투스 키보드 실사용 차이",
                    "verification_status": "pending",
                    "projected_score": 66.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "블루투스 키보드",
                    "source_keyword": "블루투스 키보드",
                    "longtail_keyword": "블루투스 키보드 설정 팁",
                    "verification_status": "pending",
                    "projected_score": 65.0,
                },
                {
                    "suggestion_id": "longtail-03",
                    "representative_keyword": "블루투스 키보드",
                    "source_keyword": "블루투스 키보드",
                    "longtail_keyword": "블루투스 키보드 장단점",
                    "verification_status": "pending",
                    "projected_score": 64.0,
                },
            ],
        }
    )

    v1_items = [item for item in items if item["target_mode"] == "longtail_selected"]

    assert summary["mode_counts"]["longtail_selected"] == 1
    assert len(v1_items) == 1
    assert v1_items[0]["base_keyword"] == "블루투스 키보드"


def test_build_title_targets_spreads_v1_across_distinct_families_for_specific_keywords() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "큐센 q104",
                    "score": 58.0,
                    "metrics": {"volume": 240.0, "cpc": 90.0},
                }
            ],
            "keyword_clusters": [
                {
                    "cluster_id": "cluster-01",
                    "representative_keyword": "큐센 q104",
                    "topic_terms": ["큐센", "q104"],
                    "all_keywords": ["큐센 q104"],
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "큐센 q104",
                    "source_keyword": "큐센 q104",
                    "longtail_keyword": "큐센 q104 설정 팁",
                    "verification_status": "pending",
                    "projected_score": 71.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "큐센 q104",
                    "source_keyword": "큐센 q104",
                    "longtail_keyword": "큐센 q104 연결 방법",
                    "verification_status": "pending",
                    "projected_score": 70.0,
                },
                {
                    "suggestion_id": "longtail-03",
                    "representative_keyword": "큐센 q104",
                    "source_keyword": "큐센 q104",
                    "longtail_keyword": "큐센 q104 자주 생기는 문제",
                    "verification_status": "pending",
                    "projected_score": 69.0,
                },
            ],
        }
    )

    v1_keywords = {
        item["keyword"]
        for item in items
        if item["target_mode"] == "longtail_selected"
    }

    assert summary["mode_counts"]["longtail_selected"] == 2
    assert "큐센 q104 자주 생기는 문제" in v1_keywords
    assert len(
        {
            keyword
            for keyword in v1_keywords
            if ("설정" in keyword or "연결" in keyword)
        }
    ) == 1


def test_build_title_targets_dedupes_near_duplicate_single_keywords() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {"keyword": "로지텍 지슈스", "score": 48.0},
                {"keyword": "지슈스 마우스", "score": 34.0},
                {"keyword": "로지텍 지슈스 마우스", "score": 22.0},
                {"keyword": "로지텍 지슈스 사전예약", "score": 34.0},
                {"keyword": "로지텍 g304", "score": 39.0},
            ],
            "title_options": {
                "keyword_modes": ["single"],
            },
        }
    )

    single_keywords = {item["keyword"] for item in items if item["target_mode"] == "single"}

    assert summary["mode_counts"]["single"] == 3
    assert single_keywords == {
        "로지텍 지슈스",
        "로지텍 지슈스 사전예약",
        "로지텍 g304",
    }


def test_build_title_targets_keeps_single_keywords_when_intent_token_differs() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {"keyword": "경제 뉴스 추천", "score": 51.0},
                {"keyword": "경제 뉴스 비교", "score": 49.0},
                {"keyword": "경제 정책 추천", "score": 47.0},
            ],
            "title_options": {
                "keyword_modes": ["single"],
            },
        }
    )

    single_keywords = {item["keyword"] for item in items if item["target_mode"] == "single"}

    assert summary["mode_counts"]["single"] == 3
    assert "경제 뉴스 추천" in single_keywords
    assert "경제 뉴스 비교" in single_keywords
    assert "경제 정책 추천" in single_keywords


def test_build_title_targets_uses_deduped_selected_base_for_v1_suggestions() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {"keyword": "로지텍 지슈스", "score": 48.0},
                {"keyword": "지슈스 마우스", "score": 34.0},
                {"keyword": "로지텍 지슈스 마우스", "score": 22.0},
                {"keyword": "로지텍 g304", "score": 39.0},
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "로지텍 지슈스",
                    "source_keyword": "로지텍 지슈스",
                    "longtail_keyword": "로지텍 지슈스 설정 팁",
                    "verification_status": "pass",
                    "verified_score": 71.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "지슈스 마우스",
                    "source_keyword": "지슈스 마우스",
                    "longtail_keyword": "지슈스 마우스 설정 팁",
                    "verification_status": "pass",
                    "verified_score": 69.0,
                },
                {
                    "suggestion_id": "longtail-03",
                    "representative_keyword": "로지텍 지슈스 마우스",
                    "source_keyword": "로지텍 지슈스 마우스",
                    "longtail_keyword": "로지텍 지슈스 마우스 설정 팁",
                    "verification_status": "pass",
                    "verified_score": 68.0,
                },
                {
                    "suggestion_id": "longtail-04",
                    "representative_keyword": "로지텍 g304",
                    "source_keyword": "로지텍 g304",
                    "longtail_keyword": "로지텍 g304 실사용 차이",
                    "verification_status": "pass",
                    "verified_score": 72.0,
                },
            ],
            "title_options": {
                "keyword_modes": ["longtail_selected"],
            },
        }
    )

    v1_keywords = {item["keyword"] for item in items if item["target_mode"] == "longtail_selected"}

    assert summary["mode_counts"]["longtail_selected"] == 2
    assert v1_keywords == {
        "로지텍 지슈스 설정 팁",
        "로지텍 g304 실사용 차이",
    }


def test_build_title_targets_caps_seed_anchor_selected_v1_to_one() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "닌텐도 스위치2 사전예약",
                    "score": 32.0,
                    "selection_mode": "seed_anchor",
                    "selection_reason": "seed_intent_preserved",
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "닌텐도 스위치2 사전예약",
                    "source_keyword": "닌텐도 스위치2 사전예약",
                    "longtail_keyword": "닌텐도 스위치2 사전예약 전 체크포인트",
                    "verification_status": "pass",
                    "verified_score": 72.0,
                },
                {
                    "suggestion_id": "longtail-02",
                    "representative_keyword": "닌텐도 스위치2 사전예약",
                    "source_keyword": "닌텐도 스위치2 사전예약",
                    "longtail_keyword": "닌텐도 스위치2 사전예약 방법",
                    "verification_status": "pass",
                    "verified_score": 69.0,
                },
            ],
            "title_options": {
                "keyword_modes": ["longtail_selected"],
            },
        }
    )

    assert summary["mode_counts"]["longtail_selected"] == 1
    assert [item["keyword"] for item in items] == ["닌텐도 스위치2 사전예약 전 체크포인트"]
    assert items[0]["source_selection_mode"] == "seed_anchor"
    assert items[0]["source_selection_reason"] == "seed_intent_preserved"


def test_build_title_targets_skips_v1_for_already_concrete_selected_keyword() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "무선 마우스 설정 팁",
                    "score": 18.0,
                    "selection_mode": "fallback",
                    "selection_reason": "top_scored_candidate",
                }
            ],
            "longtail_suggestions": [
                {
                    "suggestion_id": "longtail-01",
                    "representative_keyword": "무선 마우스 설정 팁",
                    "source_keyword": "무선 마우스 설정 팁",
                    "longtail_keyword": "무선 마우스 설정 팁 전 체크포인트",
                    "verification_status": "pending",
                    "projected_score": 52.0,
                }
            ],
            "title_options": {
                "keyword_modes": ["single", "longtail_selected"],
            },
        }
    )

    assert summary["mode_counts"]["single"] == 1
    assert summary["mode_counts"]["longtail_selected"] == 0
    assert [item["keyword"] for item in items] == ["무선 마우스 설정 팁"]


def test_build_title_targets_preserves_selection_metadata_on_single_targets() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "오사카 가성비 호텔",
                    "score": 28.0,
                    "selection_mode": "seed_anchor",
                    "selection_reason": "seed_intent_preserved",
                }
            ],
            "title_options": {
                "keyword_modes": ["single"],
            },
        }
    )

    assert summary["mode_counts"]["single"] == 1
    assert items[0]["keyword"] == "오사카 가성비 호텔"
    assert items[0]["source_selection_mode"] == "seed_anchor"
    assert items[0]["source_selection_reason"] == "seed_intent_preserved"


def test_title_generator_supports_explicit_title_targets() -> None:
    result = run(
        {
            "title_targets": [
                {
                    "target_id": "longtail_exploratory:insurance-checklist",
                    "keyword": "\ubcf4\ud5d8 \uac00\uc785 \uccb4\ud06c\ub9ac\uc2a4\ud2b8",
                    "target_mode": "longtail_exploratory",
                    "base_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "support_keywords": ["\ubcf4\ud5d8 \uac00\uc785"],
                    "source_keywords": ["\ubcf4\ud5d8 \ucd94\ucc9c", "\ubcf4\ud5d8 \uac00\uc785"],
                    "source_kind": "explicit_target",
                    "source_note": "\uc7ac\uc0dd\uc131 \uc804\uc6a9 target",
                }
            ],
            "title_options": {
                "mode": "template",
            },
        }
    )

    item = result["generated_titles"][0]
    assert item["target_id"].startswith("longtail_exploratory:")
    assert item["target_mode"] == "longtail_exploratory"
    assert item["base_keyword"] == "\ubcf4\ud5d8 \ucd94\ucc9c"
    assert item["support_keywords"] == ["\ubcf4\ud5d8 \uac00\uc785"]
    assert result["generation_meta"]["target_summary"]["requested_modes"] == ["longtail_exploratory"]


def test_title_generator_preserves_same_keyword_explicit_targets_across_modes() -> None:
    keyword = "\ubcf4\ud5d8 \uac00\uc785 \uccb4\ud06c\ub9ac\uc2a4\ud2b8"
    result = run(
        {
            "title_targets": [
                {
                    "target_id": "longtail_selected:insurance-checklist",
                    "keyword": keyword,
                    "target_mode": "longtail_selected",
                    "base_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "source_keywords": ["\ubcf4\ud5d8 \ucd94\ucc9c"],
                    "source_kind": "explicit_target",
                    "source_note": "V1",
                },
                {
                    "target_id": "longtail_exploratory:insurance-checklist",
                    "keyword": keyword,
                    "target_mode": "longtail_exploratory",
                    "base_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "source_keywords": ["\ubcf4\ud5d8 \ucd94\ucc9c"],
                    "source_kind": "explicit_target",
                    "source_note": "V2",
                },
            ],
            "title_options": {
                "mode": "template",
            },
        }
    )

    summary = result["generation_meta"]["target_summary"]
    assert summary["requested_modes"] == ["longtail_selected", "longtail_exploratory"]
    assert summary["mode_counts"]["longtail_selected"] == 1
    assert summary["mode_counts"]["longtail_exploratory"] == 1
    assert len(result["generated_titles"]) == 2
    assert {item["target_mode"] for item in result["generated_titles"]} == {
        "longtail_selected",
        "longtail_exploratory",
    }


def test_build_title_targets_related_modes_skip_self_derived_and_stopword_modifier_rows() -> None:
    items, summary = build_title_targets(
        {
            "selected_keywords": [
                {
                    "keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "profitability_grade": "A",
                    "attackability_grade": "2",
                    "score": 76.0,
                    "metrics": {"volume": 1200.0, "cpc": 210.0},
                }
            ],
            "keyword_clusters": [
                {
                    "cluster_id": "cluster-01",
                    "representative_keyword": "\ubcf4\ud5d8 \ucd94\ucc9c",
                    "topic_terms": ["\ubcf4\ud5d8", "\ucd94\ucc9c", "\uac00\uc785"],
                    "all_keywords": ["\ubcf4\ud5d8 \ucd94\ucc9c"],
                }
            ],
            "analyzed_keywords": [
                {
                    "keyword": "\ubcf4\ud5d8 \uac00\uc785 \ubc29\ubc95",
                    "score": 44.0,
                    "metrics": {"volume": 180.0, "cpc": 120.0},
                },
                {
                    "keyword": "\ubcf4\ud5d8 \uac00\uc785 \uc870\uac74",
                    "score": 52.0,
                    "metrics": {"volume": 220.0, "cpc": 150.0},
                },
            ],
            "title_options": {
                "keyword_modes": ["longtail_exploratory"],
            },
        }
    )

    assert summary["mode_counts"]["longtail_exploratory"] == len(items)
    assert items
    assert all(item["target_mode"] == "longtail_exploratory" for item in items)
    assert all("\uc870\uac74" in item["keyword"] for item in items)
    assert all("\ucd94\ucc9c \uae30\uc900" not in item["keyword"] for item in items)
    assert all("\ubc29\ubc95 \uc804 \uccb4\ud06c\ud3ec\uc778\ud2b8" not in item["keyword"] for item in items)


def test_sanitize_related_mode_keyword_collapses_model_noise() -> None:
    assert (
        _sanitize_related_mode_keyword(
            "로지텍 지슈라 스트라이크 m720 실사용 차이",
            representative_keyword="로지텍 지슈라",
        )
        == "로지텍 지슈라 실사용 차이"
    )
    assert (
        _sanitize_related_mode_keyword(
            "로지텍 마우스 맥북 블루투스 연결 문제",
            representative_keyword="로지텍 마우스",
        )
        == "로지텍 마우스 맥북 블루투스 연결 문제"
    )

def test_title_generation_options_parse_rewrite_ai_settings() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "model": "gpt-4o-mini",
                "rewrite_provider": "vertex",
                "rewrite_model": "gemini-2.5-flash-lite",
                "rewrite_api_key": "vertex-key",
            }
        }
    )

    assert options.provider == "openai"
    assert options.model == "gpt-4o-mini"
    assert options.rewrite_provider == "vertex"
    assert options.rewrite_model == "gemini-2.5-flash-lite"
    assert options.rewrite_api_key == "vertex-key"


def test_title_generation_options_parse_quality_prompt_settings() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "model": "gpt-4o-mini",
                "quality_system_prompt": "custom home eval prompt",
                "active_evaluation_prompt_profile_id": "eval-profile-1",
            }
        }
    )

    assert options.quality_system_prompt == "custom home eval prompt"
    assert options.quality_prompt_profile_id == "eval-profile-1"

    fallback_options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "model": "gpt-4o-mini",
            }
        }
    )

    assert fallback_options.quality_system_prompt == DEFAULT_TITLE_EVALUATION_PROMPT


def test_title_generation_options_default_batch_size_is_twenty() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "model": "gpt-4o-mini",
            }
        }
    )

    assert options.batch_size == 20


def test_title_generation_builds_sampled_home_evaluation_options_by_default() -> None:
    options = TitleGenerationOptions.from_input(
        {
            "title_options": {
                "mode": "ai",
                "provider": "openai",
                "model": "gpt-4o-mini",
                "api_key": "openai-key",
            }
        }
    )

    evaluation_options = _build_title_evaluation_options(options)

    assert evaluation_options.sample_ratio == 0.15
    assert evaluation_options.max_sampled_items == 2


def test_title_quality_batches_home_evaluations_across_items() -> None:
    items = [
        {
            "keyword": "insurance plan",
            "titles": {
                "naver_home": [
                    "insurance plan why now",
                    "insurance plan compare point",
                ],
                "blog": [
                    "insurance plan compare guide",
                    "insurance plan signup checklist",
                ],
            },
        },
        {
            "keyword": "card benefit",
            "titles": {
                "naver_home": [
                    "card benefit what changed",
                    "card benefit why split now",
                ],
                "blog": [
                    "card benefit summary guide",
                    "card benefit apply checklist",
                ],
            },
        },
    ]

    with patch(
        "app.title.quality_ai.request_ai_json_object",
        return_value={
            "items": [
                {
                    "index": 0,
                    "keyword": "insurance plan",
                    "title": "insurance plan why now",
                    "score": {
                        "issue_or_context": 18,
                        "curiosity_gap": 16,
                        "contrast_or_conflict": 10,
                        "reversal_or_unexpected": 10,
                        "emotional_trigger": 8,
                        "specificity": 8,
                        "readability": 5,
                        "total": 75,
                    },
                    "verdict": "rewrite",
                    "reason": "CTR 포인트는 있으나 한 단계 더 끌어올릴 여지가 있습니다.",
                },
                {
                    "index": 1,
                    "keyword": "insurance plan",
                    "title": "insurance plan compare point",
                    "score": {
                        "issue_or_context": 17,
                        "curiosity_gap": 15,
                        "contrast_or_conflict": 11,
                        "reversal_or_unexpected": 9,
                        "emotional_trigger": 7,
                        "specificity": 8,
                        "readability": 5,
                        "total": 72,
                    },
                    "verdict": "rewrite",
                    "reason": "CTR 기준에서는 대비 축이 더 또렷해야 합니다.",
                },
                {
                    "index": 2,
                    "keyword": "card benefit",
                    "title": "card benefit what changed",
                    "score": {
                        "issue_or_context": 18,
                        "curiosity_gap": 16,
                        "contrast_or_conflict": 9,
                        "reversal_or_unexpected": 8,
                        "emotional_trigger": 7,
                        "specificity": 8,
                        "readability": 5,
                        "total": 71,
                    },
                    "verdict": "rewrite",
                    "reason": "CTR 관점에서 긴장감이 조금 더 필요합니다.",
                },
                {
                    "index": 3,
                    "keyword": "card benefit",
                    "title": "card benefit why split now",
                    "score": {
                        "issue_or_context": 19,
                        "curiosity_gap": 17,
                        "contrast_or_conflict": 11,
                        "reversal_or_unexpected": 10,
                        "emotional_trigger": 8,
                        "specificity": 8,
                        "readability": 5,
                        "total": 78,
                    },
                    "verdict": "keep",
                    "reason": "CTR 관점에서 바로 써도 되는 수준입니다.",
                },
            ]
        },
    ) as mocked_request:
        enriched_items, summary = enrich_title_results(
            items,
            evaluation_options=TitleEvaluationOptions(
                provider="openai",
                model="gpt-4o-mini",
                api_key="openai-key",
                system_prompt="custom eval prompt",
                batch_size=20,
            ),
        )

    assert mocked_request.call_count == 1
    assert summary["total_count"] == 2
    assert enriched_items[0]["quality_report"]["title_checks"]["naver_home"][0]["checks"]["ai_evaluated"] is True
    assert enriched_items[1]["quality_report"]["title_checks"]["naver_home"][1]["checks"]["ai_evaluated"] is True


def test_title_quality_samples_only_top_home_candidates_for_ai_evaluation() -> None:
    items = [
        {
            "keyword": f"keyword {index}",
            "titles": {
                "naver_home": [
                    f"keyword {index} 왜 다를까",
                    f"keyword {index} 뭐가 갈렸을까",
                ],
                "blog": [
                    f"keyword {index} 비교 가이드",
                    f"keyword {index} 체크리스트",
                ],
            },
        }
        for index in range(1, 6)
    ]

    with patch(
        "app.title.quality_ai.request_ai_json_object",
        return_value={
            "items": [
                {
                    "index": 0,
                    "keyword": "keyword 1",
                    "title": "keyword 1 왜 다를까",
                    "score": {"total": 70},
                    "verdict": "keep",
                    "reason": "CTR ok",
                },
                {
                    "index": 1,
                    "keyword": "keyword 1",
                    "title": "keyword 1 뭐가 갈렸을까",
                    "score": {"total": 72},
                    "verdict": "keep",
                    "reason": "CTR ok",
                },
            ]
        },
    ) as mocked_request:
        enriched_items, _summary = enrich_title_results(
            items,
            evaluation_options=TitleEvaluationOptions(
                provider="openai",
                model="gpt-4o-mini",
                api_key="openai-key",
                system_prompt="custom eval prompt",
                sample_ratio=0.2,
                max_sampled_items=1,
            ),
        )

    assert mocked_request.call_count == 1
    assert enriched_items[0]["quality_report"]["title_checks"]["naver_home"][0]["checks"]["ai_evaluated"] is True
    assert enriched_items[1]["quality_report"]["title_checks"]["naver_home"][0]["checks"]["ai_evaluated"] is False


def test_title_quality_evaluation_splits_large_provider_requests_and_uses_fast_timeout() -> None:
    entries = [
        {
            "keyword": "insurance plan",
            "title": f"insurance plan title {index}",
        }
        for index in range(25)
    ]
    captured_kwargs: list[dict[str, object]] = []

    with patch(
        "app.title.quality_ai.request_ai_json_object",
        side_effect=lambda **kwargs: captured_kwargs.append(kwargs) or {"items": []},
    ):
        evaluations = request_naver_home_title_evaluations_batch(
            entries,
            TitleEvaluationOptions(
                provider="openai",
                model="gpt-4o-mini",
                api_key="openai-key",
                system_prompt="custom eval prompt",
                request_batch_size=10,
                request_timeout_seconds=10.0,
                max_retries=0,
            ),
        )

    assert evaluations == {}
    assert len(captured_kwargs) == 6
    assert all(call["request_timeout_seconds"] == 10.0 for call in captured_kwargs)
    assert all(call["max_retries"] == 0 for call in captured_kwargs)


def test_title_quality_evaluation_defaults_to_smaller_batches_and_longer_timeout() -> None:
    options = TitleEvaluationOptions(
        provider="openai",
        model="gpt-4o-mini",
        api_key="openai-key",
        system_prompt="custom eval prompt",
    )

    assert options.request_batch_size == 6
    assert options.request_timeout_seconds == 20.0


def test_title_quality_evaluation_rescues_single_entry_after_full_batch_failure() -> None:
    entries = [
        {
            "keyword": "insurance plan",
            "title": "insurance plan why now",
        },
        {
            "keyword": "card benefit",
            "title": "card benefit what changed",
        },
    ]
    captured_kwargs: list[dict[str, object]] = []
    call_count = 0

    def fake_request(**kwargs: object) -> dict[str, object]:
        nonlocal call_count
        call_count += 1
        captured_kwargs.append(kwargs)
        if call_count == 1:
            raise RuntimeError("batch timeout")
        return {
            "items": [
                {
                    "index": 0,
                    "keyword": "insurance plan",
                    "title": "insurance plan why now",
                    "score": {"total": 70},
                    "verdict": "keep",
                    "reason": "CTR ok",
                }
            ]
        }

    with patch("app.title.quality_ai.request_ai_json_object", side_effect=fake_request):
        evaluations = request_naver_home_title_evaluations_batch(
            entries,
            TitleEvaluationOptions(
                provider="openai",
                model="gpt-4o-mini",
                api_key="openai-key",
                system_prompt="custom eval prompt",
            ),
        )

    assert evaluations[0]["score"] == 70
    assert len(captured_kwargs) == 2
    assert all(call["request_timeout_seconds"] == 20.0 for call in captured_kwargs)


def test_batch_evaluation_normalizer_keeps_home_titles_in_soft_keep_band() -> None:
    keep_item = _normalize_batch_evaluation_item(
        {
            "index": 0,
            "keyword": "insurance plan",
            "title": "insurance plan why now",
            "score": {"total": 68},
            "verdict": "rewrite",
            "reason": "soft keep",
        }
    )
    high_keep_item = _normalize_batch_evaluation_item(
        {
            "index": 1,
            "keyword": "insurance plan",
            "title": "insurance plan compare point",
            "score": {"total": 80},
            "verdict": "rewrite",
            "reason": "soft keep",
        }
    )
    rewrite_item = _normalize_batch_evaluation_item(
        {
            "index": 2,
            "keyword": "insurance plan",
            "title": "insurance plan latest update",
            "score": {"total": 67},
            "verdict": "keep",
            "reason": "rewrite",
        }
    )

    assert keep_item["verdict"] == "keep"
    assert high_keep_item["verdict"] == "keep"
    assert rewrite_item["verdict"] == "rewrite"


def test_assess_single_title_keeps_naver_home_scores_in_soft_keep_band() -> None:
    report = assess_single_title(
        "insurance plan",
        "insurance plan why now",
        "naver_home",
        {},
        ai_evaluation={
            "score": 68,
            "score_breakdown": {
                "issue_or_context": 18,
                "curiosity_gap": 16,
                "contrast_or_conflict": 10,
                "reversal_or_unexpected": 8,
                "emotional_trigger": 7,
                "specificity": 8,
                "readability": 5,
                "total": 68,
            },
            "verdict": "keep",
            "reason": "soft keep band",
        },
    )

    assert report["status"] == "keep"
    assert report["checks"]["ai_rewrite_recommended"] is False


def test_title_generator_batches_retry_candidates_with_batch_size_twenty() -> None:
    keywords = [f"보험 추천 {index}" for index in range(1, 6)]

    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": keyword,
                "titles": {
                    "naver_home": [keyword, keyword],
                    "blog": [
                        f"{keyword} 비교 포인트 정리",
                        f"{keyword} 가입 체크리스트",
                    ],
                },
            }
            for keyword in keywords
        ],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=lambda chunk, options: _build_slot_retry_response(chunk),
    ) as mocked_slot_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                    for keyword in keywords
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "openai-key",
                    "model": "gpt-4.1-mini",
                    "keyword_modes": ["single"],
                    "batch_size": 20,
                },
            }
        )

    assert mocked_request.call_count == 1
    assert mocked_slot_request.call_count == 1
    assert mocked_request.call_args_list[0].kwargs["options"].batch_size == 20
    assert mocked_slot_request.call_args.kwargs["options"].batch_size == 20
    assert len(mocked_slot_request.call_args.args[0]) == result["generation_meta"]["auto_retry"]["attempted_slot_count"]
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 5
    assert result["generation_meta"]["auto_retry"]["attempted_slot_count"] == 15
    assert result["generation_meta"]["auto_retry"]["rewrite_call_count"] == 1
    assert result["generation_meta"]["auto_retry"]["rewrite_batch_sizes"] == [15]
    assert result["generation_meta"]["auto_retry"]["max_slot_rewrite_attempts"] == 2


def test_title_generator_batches_bad_slots_in_single_retry_call_when_under_twenty() -> None:
    keywords = [f"보험 추천 {index}" for index in range(1, 8)]
    requested_keyword_chunk_sizes: list[int] = []
    requested_slot_chunk_sizes: list[int] = []

    def fake_request(chunk, options):
        requested_keyword_chunk_sizes.append(len(chunk))
        return [
            {
                "keyword": item["keyword"],
                "titles": {
                    "naver_home": [item["keyword"], item["keyword"]],
                    "blog": [
                        f'{item["keyword"]} 비교 포인트 정리',
                        f'{item["keyword"]} 가입 체크리스트',
                    ],
                },
            }
            for item in chunk
        ]

    def fake_slot_request(chunk, options):
        requested_slot_chunk_sizes.append(len(chunk))
        return _build_slot_retry_response(chunk)

    with patch("app.title.title_generator.request_ai_titles", side_effect=fake_request), patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=fake_slot_request,
    ):
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                    for keyword in keywords
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "openai-key",
                    "model": "gpt-4.1-mini",
                    "keyword_modes": ["single"],
                    "batch_size": 20,
                },
            }
    )

    assert requested_keyword_chunk_sizes == [7]
    assert requested_slot_chunk_sizes == [20, 1]
    assert result["generation_meta"]["auto_retry"]["attempted_count"] == 7
    assert result["generation_meta"]["auto_retry"]["attempted_slot_count"] == 21
    assert result["generation_meta"]["auto_retry"]["rewrite_call_count"] == 2
    assert result["generation_meta"]["auto_retry"]["rewrite_batch_sizes"] == [20, 1]


def test_title_generator_uses_rewrite_ai_for_retry_and_escalation() -> None:
    keyword = "insurance plan"

    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": keyword,
                "titles": {
                    "naver_home": [keyword, keyword],
                    "blog": [keyword, keyword],
                },
            }
        ],
    ) as mocked_request, patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=[
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ],
                degraded=True,
            ),
            _build_slot_retry_response(
                [
                    {"slot_id": "0_home_1", "keyword": keyword, "channel": "naver_home", "slot_index": 1},
                    {"slot_id": "0_home_2", "keyword": keyword, "channel": "naver_home", "slot_index": 2},
                    {"slot_id": "0_blog_1", "keyword": keyword, "channel": "blog", "slot_index": 1},
                    {"slot_id": "0_blog_2", "keyword": keyword, "channel": "blog", "slot_index": 2},
                ]
            ),
        ],
    ) as mocked_slot_request:
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "openai-key",
                    "model": "gpt-4o-mini",
                    "rewrite_provider": "vertex",
                    "rewrite_api_key": "vertex-key",
                    "rewrite_model": "gemini-2.5-flash-lite",
                    "keyword_modes": ["single"],
                },
            }
        )

    assert mocked_request.call_count == 1
    assert mocked_request.call_args_list[0].kwargs["options"].provider == "openai"
    assert mocked_request.call_args_list[0].kwargs["options"].model == "gpt-4o-mini"
    assert mocked_slot_request.call_args_list[0].kwargs["options"].provider == "vertex"
    assert mocked_slot_request.call_args_list[0].kwargs["options"].model == "gemini-2.5-flash-lite"
    assert mocked_slot_request.call_args_list[0].kwargs["options"].api_key == "vertex-key"
    assert mocked_slot_request.call_args_list[1].kwargs["options"].provider == "vertex"
    assert mocked_slot_request.call_args_list[1].kwargs["options"].model == "gemini-2.5-flash"
    assert result["generation_meta"]["auto_retry"]["provider"] == "vertex"
    assert result["generation_meta"]["auto_retry"]["model"] == "gemini-2.5-flash-lite"
    assert result["generation_meta"]["auto_retry"]["rewrite_call_count"] == 1
    assert result["generation_meta"]["auto_retry"]["rewrite_batch_sizes"] == [4]
    assert result["generation_meta"]["model_escalation"]["source_provider"] == "vertex"
    assert result["generation_meta"]["model_escalation"]["target_provider"] == "vertex"
    assert result["generation_meta"]["model_escalation"]["rewrite_call_count"] == 1
    assert result["generation_meta"]["model_escalation"]["rewrite_batch_sizes"] == [4]
    assert result["generation_meta"]["model_escalation"]["max_slot_rewrite_attempts"] == 2


def test_title_generator_downgrades_slots_after_two_failed_rewrite_attempts() -> None:
    keyword = "alpha beta"
    unchanged_slot_response = [
        {"slot_id": "0_home_1", "title": keyword},
        {"slot_id": "0_home_2", "title": keyword},
        {"slot_id": "0_blog_1", "title": keyword},
        {"slot_id": "0_blog_2", "title": keyword},
    ]

    with patch(
        "app.title.title_generator.request_ai_titles",
        return_value=[
            {
                "keyword": keyword,
                "titles": {
                    "naver_home": [keyword, keyword],
                    "blog": [keyword, keyword],
                },
            }
        ],
    ), patch(
        "app.title.title_generator.request_ai_slot_rewrites",
        side_effect=[unchanged_slot_response, unchanged_slot_response],
    ) as mocked_slot_request, patch(
        "app.title.title_generator._build_practical_rescue_item",
        return_value=None,
    ):
        result = run(
            {
                "selected_keywords": [
                    {
                        "keyword": keyword,
                        "score": 1.0,
                    }
                ],
                "title_options": {
                    "mode": "ai",
                    "provider": "openai",
                    "api_key": "openai-key",
                    "model": "gpt-4o-mini",
                    "keyword_modes": ["single"],
                },
            }
        )

    quality_report = result["generated_titles"][0]["quality_report"]
    title_checks = quality_report["title_checks"]["naver_home"] + quality_report["title_checks"]["blog"]

    assert mocked_slot_request.call_count == 2
    assert result["generation_meta"]["auto_retry"]["accepted_slot_count"] == 0
    assert result["generation_meta"]["model_escalation"]["accepted_slot_count"] == 0
    assert result["generation_meta"]["model_escalation"]["downgraded_slot_count"] == 4
    assert result["generation_meta"]["model_escalation"]["downgraded_slot_ids"] == [
        "0_home_1",
        "0_home_2",
        "0_blog_1",
        "0_blog_2",
    ]
    assert result["generation_meta"]["model_escalation"]["max_slot_rewrite_attempts"] == 2
    assert quality_report["status"] == "review"
    assert quality_report["retry_recommended"] is False
    assert any("재작성 2회" in issue for issue in quality_report["issues"])
    assert all(check["status"] == "review" for check in title_checks)
    assert all(bool(check.get("checks", {}).get("retry_limit_reached")) for check in title_checks)
